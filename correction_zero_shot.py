'''This code is designed for error correction tasks under ZERO-SHOT conditions. 
Compared to the regular ZeroEC version, this version will not use automatic 
chain-of-thought generation or correction rule generation.
Instead, it solely relies on the retrieved relevant tuples as context and a
small number of general correction examples as guidance for few-shot prompting and retrieval-augmented generation.'''
import os
import time
import json
import re
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.metrics import mutual_info_score
from sklearn.cluster import KMeans, MiniBatchKMeans
from fast_sentence_transformers import FastSentenceTransformer
from langchain.prompts import (
    ChatPromptTemplate,
    HumanMessagePromptTemplate,
    SystemMessagePromptTemplate,
)
from langchain_openai import ChatOpenAI
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain_core.output_parsers import JsonOutputParser
from langchain_core.embeddings import Embeddings
from langchain_community.vectorstores import FAISS
from langchain_community.vectorstores.utils import DistanceStrategy
import ast
import copy
import multiprocessing
from collections import Counter


def get_folder_name(base_path):
    # Get the save path for the current run with date and time
    if not os.path.exists(base_path):
        print(f'base_path:{base_path} does not exist! Trying to create a new one.\n')
        os.makedirs(base_path)
        print(f'base_path:{base_path} created!\n')
    existing_folders = [d for d in os.listdir(base_path) if
                        os.path.isdir(os.path.join(base_path, d)) and d.startswith('run-')]
    max_run = 0
    for folder in existing_folders:
        run_number = int(folder.split('-')[1].split('_')[0])  # Extracting the run number before the date
        if run_number > max_run:
            max_run = run_number
    from datetime import datetime
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M")  # Format: YYYY-MM-DD_HH-MM
    next_run_folder = f'run-{max_run + 1}_{current_datetime}'
    next_run_path = os.path.join(base_path, next_run_folder)
    os.makedirs(next_run_path)
    return next_run_path


def load_prompts(*file_paths: str) -> tuple:
    return tuple(open(path, 'r', encoding='utf-8').read() for path in file_paths)


def load_examples(file_path: str) -> str:
    with open(file_path, 'r', encoding='utf-8') as file:
        content = ast.literal_eval(file.read())
    return form_examples(content)


class myEmbeddings(Embeddings):
    # Custom embedding
    def __init__(self, modelPath):
        self.model = FastSentenceTransformer(modelPath, device="cuda", quantize=True)

    def embed_documents(self, texts: List[str]) -> np.ndarray:
        return self.model.encode(texts, normalize_embeddings=True,batch_size=64,show_progress_bar=True,device="cuda:0,cuda:1,cuda:2,cuda:3")

    def embed_query(self, text: str) -> np.ndarray:
        return self.model.encode(text, normalize_embeddings=True)


def embed_data(dirty_data, embeddingModel):
    elements_list = dirty_data.values.flatten().tolist()
    embeddings = embeddingModel.embed_documents(elements_list)
    embedding_dimension = len(embeddings[0])
    embeddings_matrix = embeddings.reshape(row_count, column_count, embedding_dimension)
    return embeddings_matrix, embedding_dimension


def calc_stasts(dirty_data):
    # Calculate the statistics of the dirty data
    # 1. value frequency
    # Calculate value frequency for each element in the dataframe
    # Value frequency is defined as the count of occurrences of a value in its column divided by the total count of elements in that column
    # Create a numpy array with the same shape as dirty_data to store value frequencies
    value_freq_start_time = time.time()
    value_frequencies = np.zeros_like(dirty_data, dtype=float)
    # Get the total number of rows
    column_counts = dirty_data.shape[0]
    
    # Calculate value frequency for each column
    for col_idx, column in enumerate(dirty_data.columns):
        # Count occurrences of each value in the column
        value_counts = dirty_data[column].value_counts()
        
        # Fill the frequency array for this column
        for row_idx in range(len(dirty_data)):
            value = dirty_data.iloc[row_idx, col_idx]
            # Get frequency of this value in its column
            if value in value_counts:
                frequency = value_counts[value] / column_counts
                value_frequencies[row_idx, col_idx] = frequency
    value_freq_end_time = time.time()
    print(f"Value frequency calculation time: {value_freq_end_time - value_freq_start_time} seconds")

    # 2. vicinity frequency
    # Calculate vicinity frequency for each element in the dataframe
    # Vicinity frequency is defined as the co-occurrence count of a value with other values in the same tuple
    # divided by the occurrence count of those other values
    vicinity_freq_start_time = time.time()
    # Create a 3D numpy array to store vicinity frequencies
    # Shape: (rows, columns, columns) where the third dimension represents vicinity frequencies with each column
    row_count, column_count = dirty_data.shape
    vicinity_frequencies = np.zeros((row_count, column_count, column_count), dtype=float)
    
    # Convert DataFrame to numpy array for faster access
    dirty_data_array = dirty_data.values
    
    # Pre-compute value counts for each column - use numpy for speed
    column_value_counts = {}
    for col_idx in range(column_count):
        unique_values, counts = np.unique(dirty_data_array[:, col_idx], return_counts=True)
        column_value_counts[col_idx] = dict(zip(unique_values, counts))
    
    # Calculate co-occurrences in a vectorized way
    for attr_idx in range(column_count):
        for other_attr_idx in range(column_count):
            if attr_idx == other_attr_idx:
                continue  # Skip same column comparisons
                
            # Create a unique tuple representation for each row's value pair
            value_pairs = np.array([(dirty_data_array[i, attr_idx], dirty_data_array[i, other_attr_idx]) 
                                   for i in range(row_count)])
            
            # Count co-occurrences efficiently
            unique_pairs, pair_counts = np.unique(value_pairs, axis=0, return_counts=True)
            pair_dict = {(pair[0], pair[1]): count for pair, count in zip(unique_pairs, pair_counts)}
            
            # Fill vicinity frequencies in one go
            for tuple_idx in range(row_count):
                current_value = dirty_data_array[tuple_idx, attr_idx]
                other_value = dirty_data_array[tuple_idx, other_attr_idx]
                
                other_value_count = column_value_counts[other_attr_idx].get(other_value, 0)
                if other_value_count > 0:
                    co_occurrence = pair_dict.get((current_value, other_value), 0)
                    vicinity_frequencies[tuple_idx, attr_idx, other_attr_idx] = co_occurrence / other_value_count
    
    vicinity_freq_end_time = time.time()
    print(f"Vicinity frequency calculation time: {vicinity_freq_end_time - vicinity_freq_start_time} seconds")

    # 3. pattern frequency
    pattern_freq_start_time = time.time()
    # Create a 3D numpy array to store pattern frequencies
    # Shape: (rows, columns, 3) where the third dimension represents the three pattern types
    pattern_frequencies = np.zeros((row_count, column_count, 3), dtype=float)
    
    # Vectorized functions for pattern extraction
    def get_l1_pattern(values):
        return np.array([len(str(v)) for v in values])
    
    def get_l2_pattern(values):
        patterns = []
        for v in values:
            v = str(v)
            letter_count = sum(c.isalpha() for c in v)
            digit_count = sum(c.isdigit() for c in v)
            other_count = len(v) - letter_count - digit_count
            patterns.append((letter_count, digit_count, other_count))
        return patterns
    
    def get_l3_pattern(values):
        patterns = []
        for v in values:
            v = str(v)
            upper_count = sum(c.isupper() for c in v)
            lower_count = sum(c.islower() for c in v)
            digit_count = sum(c.isdigit() for c in v)
            other_count = len(v) - upper_count - lower_count - digit_count
            patterns.append((upper_count, lower_count, digit_count, other_count))
        return patterns
    
    # Process each column once to extract patterns and calculate frequencies
    for col_idx in range(column_count):
        column_values = dirty_data.iloc[:, col_idx].astype(str).values
        
        # Extract patterns for all values in the column at once
        l1_patterns_all = get_l1_pattern(column_values)
        l2_patterns_all = get_l2_pattern(column_values)
        l3_patterns_all = get_l3_pattern(column_values)
        
        # Count pattern occurrences using Counter

        l1_counter = Counter(l1_patterns_all)
        l2_counter = Counter(map(tuple, l2_patterns_all))
        l3_counter = Counter(map(tuple, l3_patterns_all))
        
        # Calculate frequencies for each row in this column
        for row_idx in range(row_count):
            # L1 pattern frequency
            l1_pattern = l1_patterns_all[row_idx]
            pattern_frequencies[row_idx, col_idx, 0] = l1_counter[l1_pattern] / row_count
            
            # L2 pattern frequency
            l2_pattern = l2_patterns_all[row_idx]
            pattern_frequencies[row_idx, col_idx, 1] = l2_counter[l2_pattern] / row_count
            
            # L3 pattern frequency
            l3_pattern = l3_patterns_all[row_idx]
            pattern_frequencies[row_idx, col_idx, 2] = l3_counter[l3_pattern] / row_count
    
    pattern_freq_end_time = time.time()
    print(f"Pattern frequency calculation time: {pattern_freq_end_time - pattern_freq_start_time} seconds")
    
    # 4. Concatenate all features for each element in dirty_data
    combine_features_start_time = time.time()
    # For each element, we need to concatenate:
    # - value frequency (1 value)
    # - vicinity frequency (column_count values)
    # - pattern frequency (3 values)
    # Total feature dimension: 1 + column_count + 3 = 4 + column_count
    feature_dimension = 4 + column_count
    combined_features = np.zeros((row_count, column_count, feature_dimension), dtype=float)
    
    for row_idx in range(row_count):
        for col_idx in range(column_count):
            # Add value frequency (1 value)
            combined_features[row_idx, col_idx, 0] = value_frequencies[row_idx, col_idx]
            
            # Add vicinity frequencies (column_count values)
            combined_features[row_idx, col_idx, 1:column_count+1] = vicinity_frequencies[row_idx, col_idx, :]
            
            # Add pattern frequencies (3 values)
            combined_features[row_idx, col_idx, column_count+1:] = pattern_frequencies[row_idx, col_idx, :]
    combine_features_end_time = time.time()
    print(f"Feature combination time: {combine_features_end_time - combine_features_start_time} seconds")
    
    total_time = combine_features_end_time - value_freq_start_time
    print(f"Total statistics calculation time: {total_time} seconds")
    
    return combined_features
    

def initialize_llm(model_name, openai_api_base, openai_api_key, temperature):
    return ChatOpenAI(
        model_name = model_name,
        openai_api_base = openai_api_base,
        openai_api_key = openai_api_key,
        temperature = temperature
    )


class Output(BaseModel):
    # The output parser
    chain_of_thought_for_correction: str = Field(...,
                                                 description="The chain_of_thought that led to the proposed correction")
    correction: dict = Field(...,
                             description="the most probable correction for the dirty value")


def form_examples(examps) -> str:
    # Accepts an examples data and returns a string for the prompt
    # examps is an array where each element is a dictionary containing 'input' and 'output'
    few_shot_examps_str = ''
    for examp in examps:
        examp_str = 'human: ' + \
                    json.dumps(examp['input']) + \
                    '\n' + \
                    'ai: ' + \
                    json.dumps(examp['output']) + \
                    '\n'
        few_shot_examps_str = few_shot_examps_str + examp_str
    return few_shot_examps_str


def select_representative_tuples(embeddings_matrix: np.ndarray, stats_features: np.ndarray, detection: pd.DataFrame, num_clusters: int) -> list:
    # Reshape embeddings
    m, n, l = embeddings_matrix.shape
    reshaped_embeddings = embeddings_matrix.reshape((m, n * l))
    
    # Reshape stats features
    s1, s2, s3 = stats_features.shape
    reshaped_stats = stats_features.reshape((s1, s2 * s3))
    
    # Combine embeddings and stats features directly without filtering
    combined_features = np.concatenate((reshaped_embeddings, reshaped_stats), axis=1)

    # Store original indices (all rows)
    original_indices = np.arange(m)
    
    # Perform clustering on combined features
    kmeans = KMeans(n_clusters=num_clusters, random_state=99)
    clusters = kmeans.fit_predict(combined_features)

    # Greedy selection of candidates from each cluster
    selected_indices = []
    covered_rows = set()  # Track covered rows
    covered_columns = set()  # Track covered columns with errors

    for i in range(num_clusters):
        cluster_mask = clusters == i
        if np.any(cluster_mask):
            cluster_detection = detection.iloc[cluster_mask]
            cluster_original_indices = original_indices[cluster_mask]

            # Greedy selection within cluster
            best_index = None
            best_score = -1

            for idx, orig_idx in zip(cluster_detection.index, cluster_original_indices):
                row = cluster_detection.loc[idx]
                
                # Count new error columns this row would cover
                new_error_columns = set(row[row == 1].index) - covered_columns
                
                # Count total errors in this row
                total_errors_in_row = row.sum()
                
                # Check if this row is already covered
                row_already_covered = idx in covered_rows
                
                # Calculate score: prioritize rows with more errors and new column coverage
                # Higher weight for new column coverage to ensure broader coverage
                score = (len(new_error_columns) * 2) + (total_errors_in_row * 1) - (row_already_covered * 0.5)
                
                if score > best_score:
                    best_score = score
                    best_index = idx
                    best_orig_index = orig_idx

            if best_index is not None:
                selected_indices.append(best_orig_index)
                covered_rows.add(best_index)
                # Update covered columns with the newly covered error columns
                covered_columns.update(
                    set(cluster_detection.loc[best_index][cluster_detection.loc[best_index] == 1].index))

    return selected_indices


def select_representative_tuples_old(embeddings_matrix: np.ndarray, detection: pd.DataFrame, num_clusters: int) -> list:
    # Select rows to repair based on detection
    mask_rows = detection.sum(axis=1) > 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    filtered_detection = detection[mask_rows]

    # Store original indices
    original_indices = np.where(mask_rows)[0]

    # Reshape embeddings
    m, n, l = filtered_embeddings.shape
    mask = filtered_detection.astype(bool)
    masked_embeddings = filtered_embeddings * mask.values[..., np.newaxis]
    reshaped_embeddings = masked_embeddings.reshape((m, n * l))

    # Perform clustering
    kmeans = MiniBatchKMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)

    # Greedy selection of candidates from each cluster
    selected_indices = []
    covered_columns = set()

    for i in range(num_clusters):
        cluster_mask = clusters == i
        if np.any(cluster_mask):
            cluster_detection = filtered_detection.iloc[cluster_mask]
            cluster_original_indices = original_indices[cluster_mask]

            # Greedy selection within cluster
            best_index = None
            best_new_coverage = -1

            for idx, orig_idx in zip(cluster_detection.index, cluster_original_indices):
                row = cluster_detection.loc[idx]
                new_columns = set(row[row == 1].index) - covered_columns
                if len(new_columns) > best_new_coverage:
                    best_new_coverage = len(new_columns)
                    best_index = idx
                    best_orig_index = orig_idx

            if best_index is not None:
                selected_indices.append(best_orig_index)
                covered_columns.update(
                    set(cluster_detection.loc[best_index][cluster_detection.loc[best_index] == 1].index))

    return selected_indices


def calc_mi_2(df, target_column):
    # Calculate mutual information
    # Make sure the target_column exists in dataframe
    from joblib import Parallel, delayed
    
    def calculate_mi_optimized(column, target_column, df):
        if column == target_column:
            return mutual_info_score(df[target_column], df[column])
        else:
            # Calculate the count of each combination and filter out combinations that only appear once
            count = df.groupby([target_column, column])[target_column].transform('size')
            filtered_df = df[count >= 2]
            # If the filtered data is empty, return 0
            if filtered_df.empty:
                return 0
            # Rebuild target and column to calculate MI
            A=filtered_df[target_column].reset_index(drop=True)
            B=filtered_df[column].reset_index(drop=True)
            return mutual_info_score(A, B)

    # Explicitly set number of jobs instead of using n_jobs=-1
    try:
        num_cores = multiprocessing.cpu_count()
    except:
        num_cores = 1  # Fallback to single core if detection fails

    # Calculate mutual information for all columns in parallel
    mutual_info_list = Parallel(n_jobs=num_cores)(
        delayed(calculate_mi_optimized)(col, target_column, df) for col in df.columns
    )
    # Find the maximum mutual information value and normalize
    max_mutual_info = max(mutual_info_list) if mutual_info_list else 0
    normalized_mi = [mi / max_mutual_info if max_mutual_info != 0 else 0 for mi in mutual_info_list]
    return normalized_mi


def build_retriever(dirty_data, detection, vectors_matrix, target_column):
    # Record the time when the code starts executing
    start_time = time.time()

    # 1. Convert texts
    texts = dirty_data.values.tolist()
    # print(f"Step 1: Converting dirty data to texts - Time Taken: {time.time() - start_time:.4f}s")

    # 2. Calculate mutual information
    step_start_time = time.time()
    normMI = calc_mi_2(dirty_data, target_column)
    normMI = np.round(normMI, decimals=1)
    # print(f"Step 2: Calculating Mutual Information (MI) - Time Taken: {time.time() - step_start_time:.4f}s")
    
    # print(normMI)
    normMI = np.array(normMI)
    threshold = 0.5

    # 3. Filter relevant columns
    step_start_time = time.time()
    indices = np.where(normMI >= threshold)[0]
    CoE = normMI[indices]
    embeddings_matrix_col_filtered = vectors_matrix[:, indices, :]
    texts_col_filtered = [[row[col] for col in indices] for row in texts]
    detection_col_filtered = detection.iloc[:, indices]
    # print(f"Step 3: Filtering columns based on MI - Time Taken: {time.time() - step_start_time:.4f}s")

    # 4. Merge texts
    formatted_rows = []
    header_col_filtered = detection_col_filtered.columns
    for text_row, detection_row in zip(texts_col_filtered, detection_col_filtered.values.tolist()):
        formatted_rows.append(format_row_2(text_row, header_col_filtered, detection_row))
    # print(f"Step 4: Merging texts with formatted detection - Time Taken: {time.time() - step_start_time:.4f}s")

    # 5. Reshape embeddings_matrix
    embeddings_matrix_col_row_filtered_reshaped = embeddings_matrix_col_filtered.reshape(
        embeddings_matrix_col_filtered.shape[0], -1)
    # print(f"Step 5: Reshaping embeddings matrix - Time Taken: {time.time() - step_start_time:.4f}s")

    # 6. Generate paired data
    # Use a generator expression to avoid occupying too much memory at once
    paired_data_generator = ((text_row, vector) for text_row, vector in zip(formatted_rows, embeddings_matrix_col_row_filtered_reshaped))

    # If you need to convert it to a list later, you can do so as needed
    paired_data = list(paired_data_generator)

    # print(f"Step 6: Creating paired data - Time Taken: {time.time() - step_start_time:.4f}s")

    # 7. Generate IDs
    ids = [str(i) for i in range(len(dirty_data))]
    # print(f"Step 7: Generating IDs - Time Taken: {time.time() - step_start_time:.4f}s")

    # 8. Generate meta_data
    meta_data = [{'index': i} for i in range(len(dirty_data))]
    # print(f"Step 8: Generating meta data - Time Taken: {time.time() - step_start_time:.4f}s")

    # 9. Create FAISS index
    db = FAISS.from_embeddings(text_embeddings=paired_data,
                               embedding=embeddingModel,
                               metadatas=meta_data,
                               ids=ids,
                               distance_strategy=DistanceStrategy.DOT_PRODUCT)
    retriever = db
    # print(f"Step 9: Creating FAISS index - Time Taken: {time.time() - step_start_time:.4f}s")

    # Output the total execution time
    # print(f"Total Time Taken: {time.time() - start_time:.4f}s")
    return retriever, indices, CoE


# def update_retriever(column):
#     # Update the retriever using the user's repaired results
#     ids = [str(i) for i in representative_tuples_list]
#     retriever_dict[column].delete(ids)
#     embeddings_matrix_only_repaired_col_filtered = embeddings_matrix_only_repaired[:, indices_dict[column], :]
#     texts = dirty_data_only_repaired.values.tolist()
#     texts_col_filtered = [[row[col] for col in indices_dict[column]] for row in texts]
#     header_col_filtered = [header[i] for i in indices_dict[column]]
#     formatted_rows = []
#     for text_row in texts_col_filtered:
#         formatted_rows.append(format_row(text_row, header_col_filtered))
#     meta_data = [{'index': i} for i in representative_tuples_list]
#     paired_data = [(text, vector.flatten().tolist()) for text, vector in
#                    zip(formatted_rows, embeddings_matrix_only_repaired_col_filtered)]
#     retriever_dict[column].add_embeddings(text_embeddings=paired_data,
#                                           metadatas=meta_data,
#                                           ids=ids
#                                           )


def sort_dicts(dict_list, key1, key2, key3):
    return sorted(dict_list, key=lambda x: (x.get(key1, float('inf')),
                                            x.get(key2, float('inf')),
                                            x.get(key3, float('inf'))))


# def get_auto_cot(repair_list, column, retriever, CoE, indices, detection_filtered, rep_error_info: dict, sp_examps):
#     rep_error_info[column] = {}
#     human_message_small_template = HumanMessagePromptTemplate.from_template(human_CoT_small)
#     human_input = '['
#     relevant_rows_list = []
#     dirty_tuples = []
#     dirty_tuples_list = []
#     dirty_values = []
#     clean_values = []
#     filtered_header = [header[i] for i in indices]
#     column_detection = detection_human_repaired[column].values
#     column_indices = indices_dict[column]
#     column_sums = detection_human_repaired.iloc[:, column_indices].values.sum(axis=1)
#     for row_idx in repair_list:
#         if detection.loc[row_idx, column] == 1:
#             dirty_value = dirty_data.loc[row_idx, column]
#             dirty_values.append(dirty_value)
#             clean_value = clean_data.loc[row_idx, column]
#             clean_values.append(clean_value)
#             dirty_row = dirty_data.iloc[:, indices].loc[row_idx].tolist()
#             dirty_tuples_list.append(dirty_row)
#             dirty_tuples.append(dirty_data.iloc[:, indices].loc[row_idx])
#             relevant_clean_tuples = ''
#             embeddings_row = embeddings_matrix[row_idx]
#             embeddings_row_filtered = embeddings_row[indices]
#             for i in range(len(embeddings_row_filtered)):
#                 if detection_filtered.iloc[row_idx, i] == 1:
#                     embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
#             embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
#             embeddings_row_united = embeddings_row_filtered.flatten()
#             relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
#                                                                              k=30)
#             relevant_rows_dict_list = [
#                 {
#                     'page_content': row[0].page_content,
#                     'index': idx,
#                     'score': round(row[1], 2),
#                     'target_column': column_detection[idx],
#                     'sum': column_sums[idx]
#                 }
#                 for row in relevant_rows
#                 for idx in [row[0].metadata['index']]
#             ]

#             sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
#                                                         'sum')
#             for row in sorted_relevant_rows_dict_list[:3]:
#                 relevant_clean_tuples += row['page_content'] + '\n'
#             relevant_rows_list.append(relevant_clean_tuples)
#             human_input += '{' + human_message_small_template.format(
#                 Dirty_Tuple = format_row(dirty_row, filtered_header),
#                 Erroneous_value = '{' + f'"{column}": "{dirty_value}"' + '}',
#                 Relevant_clean_tuples = relevant_clean_tuples,
#                 Correction = clean_value
#             ).content + '},'
#     human_input += ']'
#     prompt_auto_cot = ChatPromptTemplate(messages=[
#         SystemMessagePromptTemplate.from_template(sys_CoT),
#         HumanMessagePromptTemplate.from_template(human_CoT_large),
#     ],
#         partial_variables={
#             "examples": general_examples_CoT_str,
#         }
#     )
#     chain_auto_cot = (prompt_auto_cot | LLM4CoT)
#     while True: 
#         try:
#             repair_result = chain_auto_cot.invoke(
#                 {"human_input": human_input}).content
#             repair_result = repair_result.replace('\n', '')
#             repair_result = re.sub(r'```json', '', repair_result)
#             repair_result = re.sub(r'```', '', repair_result)
#             repair_result_list = ast.literal_eval(repair_result)
#             break
#         except Exception as e:
#             print('CoT failed', e)
    
#     # Build examples
#     specific_examples_llm = []
#     for idx, result in enumerate(repair_result_list):
#         dirty_tuple = dirty_tuples_list[idx]
#         json_string = """
#                             {
#                                 "input": {
#                                     "Dirty Data Tuple": null,
#                                     "Erroneous Value": null,
#                                     "Relevant Clean Tuples": null
#                                 },
#                                 "output": {
#                                     "chain_of_thought_for_correction": null,
#                                     "correction": null
#                                 }
#                             }
#                       """
#         d = json.loads(json_string)
#         d['input']['Dirty Data Tuple'] = format_row(dirty_tuple, filtered_header)
#         d['input'][
#             'Erroneous Value'] = '{' + f'"{column}": "{dirty_values[idx]}"' + '}'
#         d['input']['Relevant Clean Tuples'] = relevant_rows_list[idx],
#         d['output']['chain_of_thought_for_correction'] = result['chain_of_thought_for_correction']
#         d['output']['correction'] = '{"' + column + '\": ' + '"' + clean_values[idx] + '"}'
#         rep_error_info[column][dirty_tuples[idx].name] = {}
#         rep_error_info[column][dirty_tuples[idx].name]['dirty_tuple'] = dirty_tuples[idx].to_dict()
#         rep_error_info[column][dirty_tuples[idx].name]['dirty_value'] = dirty_values[idx]
#         rep_error_info[column][dirty_tuples[idx].name]['ground_truth'] = clean_values[idx]
#         rep_error_info[column][dirty_tuples[idx].name]['error_analysis'] = result['chain_of_thought_for_correction']
#         rep_error_info[column][dirty_tuples[idx].name]['error_type'] = result['error_type']
#         specific_examples_llm.append(d)
#     few_shot_specific_str = form_examples(specific_examples_llm)
#     sp_examps[column] = few_shot_specific_str


def sel_clean(num_clusters:int):
    # Cluster the correct tuples and randomly select one from each cluster
    # Create a mask to select rows with zero errors
    mask_rows = detection.sum(axis=1) == 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    
    # Reshape embeddings for clustering
    m, n, l = filtered_embeddings.shape
    reshaped_embeddings = filtered_embeddings.reshape(m, n * l)
    if m == 0:
        return pd.DataFrame()
    if m < num_clusters:
        num_clusters = m
    # Perform KMeans clustering
    kmeans = MiniBatchKMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)

    # Get original indices of the selected clusters
    original_indices = np.where(mask_rows)[0]
    
    # Randomly select one sample from each cluster
    selected_indices = [
        original_indices[np.random.choice(np.where(clusters == cluster_id)[0])]
        for cluster_id in range(num_clusters)
        if np.any(clusters == cluster_id)
    ]
    
    return dirty_data.iloc[selected_indices]


def train_val_split(data:dict): 
    train_data = {}
    val_data = {}
    keys = list(data.keys())
    
    # Split the data with a fixed random seed
    random.seed(0)
    random.shuffle(keys)
    split_ratio = 0.5
    # Calculate the split index
    split_index = int(len(data) * split_ratio + 0.5)
    # Split the key list
    train_keys = keys[:split_index]
    val_keys = keys[split_index:]
    # Split the data
    train_data = {key: data[key] for key in train_keys}
    val_data = {key: data[key] for key in val_keys}
    return train_data, val_data 


def clean_data_integration(clean_data:pd.DataFrame, rep_data_info:dict):
    # For each error column, merge the sampled clean_data with dirty_data
    for column in dirty_data.columns:
         if detection[column].sum() > 0:
             for idx in range(len(clean_data)):
                rep_data_info[column][clean_data.iloc[idx].name] = {}
                rep_data_info[column][clean_data.iloc[idx].name]['dirty_tuple'] = clean_data.iloc[idx].to_dict()
                rep_data_info[column][clean_data.iloc[idx].name]['dirty_value'] = clean_data.iloc[idx][column]
                rep_data_info[column][clean_data.iloc[idx].name]['error_analysis'] = 'This is a clean value that does not need correction.'
                rep_data_info[column][clean_data.iloc[idx].name]['ground_truth'] = clean_data.iloc[idx][column]
                rep_data_info[column][clean_data.iloc[idx].name]['error_type'] = 'clean'
    pass


# def code_generation(train_data:dict,column:str,codes:dict):
#     # If it is a formatting issue, let the LLM generate code
#     # indices = indices_dict[column]
#     # filtered_header = [header[i] for i in indices]
#     examples = ''
#     for error in train_data.values():
#         dirty_value_str = error['dirty_value']
#         analysis = error['error_analysis']
#         clean_value_str = error['ground_truth']
#         examples += f"[Erroneous_value: {dirty_value_str}\n"
#         examples += f"Error_analysis: {analysis}\n"
#         examples += f"Correct_value: {clean_value_str}]\n"
#     prompt_code_generation = ChatPromptTemplate(messages=[
#         SystemMessagePromptTemplate.from_template(sys_code_generation),
#         HumanMessagePromptTemplate.from_template(human_code_generation),
#         ])
#     chain_code_generation = (prompt_code_generation | LLM4Code)
#     while True:
#         try:
#             if column == 'ounces':
#                 prompt = prompt_code_generation.invoke({'examples': examples})
#                 print(prompt.to_string())
#             raw_code = chain_code_generation.invoke({'examples': examples}).content
#             extracted_code = re.findall(r'```python(.*?)```', raw_code, re.DOTALL)
#             if extracted_code:
#                 codes[column] = extracted_code[0]
#             break
#         except Exception as e:
#             print('code generation failed', e)


def correct():
    pass


# def FD_generation(train_data:dict,column:str,FDs:dict):
#     # For VAD, MISS, TYPO, find potential FD to complete the repair
#     indices = indices_dict[column]
#     filtered_header = [header[i] for i in indices]
#     examples = ''
#     for error in train_data.values():
#         dirty_tuple_str = json.dumps(error['dirty_tuple'])
#         dirty_value_str = f"{column}: {error['dirty_value']}"
#         analysis = error['error_analysis']
#         correct_value = f"{column}: {error['ground_truth']}"
#         examples += f"Dirty_tuple: {dirty_tuple_str}\n"
#         examples += f"Erroneous_value: {dirty_value_str}, "
#         examples += f"Error_analysis: {analysis}\n"
#         examples += f"Correct_value: {correct_value}]\n"
#     prompt_fd_generation = ChatPromptTemplate(messages=[
#         SystemMessagePromptTemplate.from_template(sys_fd_generation),
#         ])
#     chain_fd_generation = (prompt_fd_generation | LLM4FD)
#     while True:
#         try:
#             fd_raw = chain_fd_generation.invoke({'examples': examples}).content
#             fd_raw = re.sub(r'```json', '', fd_raw)
#             fd_raw = re.sub(r'```', '', fd_raw)
#             # pattern = r'\{[^{}]*\}'
#             # fd_raw = re.findall(pattern, fd_raw)
#             fd_dict = ast.literal_eval(fd_raw)
#             break
#         except Exception as e:
#             print('fd generation failed', e)
#     FDs[column] = fd_dict
#     pass


# def code_evaluation_execution(code, val_data, column):
#     # Run code on val_data to check if it can fix errors
#     # Use exec() to execute code
#     exec(code, globals())
#     # Check the repair results
#     flag=True
#     for error in val_data.values():
#         if str(correct(error['dirty_value'])) != str(error['ground_truth']):
#             print(f"expect {error['ground_truth']}, get {correct(error['dirty_value'])}")
#             flag = False
#     if flag==False:
#         print(f'{column} code failed')
#         return flag
#     # If passed, execute code on corrections
#     mask = detection_human_repaired[column] == 1
#     corrections.loc[mask, column] = corrections.loc[mask, column].apply(correct)
#     corrections[column] = corrections[column].astype(str)
#     print(f'{column} code passed')
#     return True


# def fd_evaluation_execution(fd:dict, val_data:dict, column:str):
#     match = re.search(r"^(.+?)\s*→\s*(.+)$", fd['functional_dependency'])
#     if match:
#         attr1, attr2 = match.groups()
#     else:
#         return False
#     # Verify fd on val_data
#     # When verifying clean_data, exclude yourself
#     if attr1 == 'None':
#         return False
#     flag = True
#     filter_conditions = (detection_human_repaired[attr1] == 0) & (detection_human_repaired[attr2] == 0)

#     for idx, error in val_data.items():
#         if error['error_type'] == 'clean':
#             filter_conditions &= (dirty_data_human_repaired.index != idx)
    
#         if attr1 in error['dirty_tuple']:
#             valid_rows = dirty_data_human_repaired[attr1] == error['dirty_tuple'][attr1]
#         else:
#             valid_rows = pd.Series([False] * len(dirty_data_human_repaired))
            
#         attr2_values = dirty_data_human_repaired.loc[valid_rows & filter_conditions, attr2]
        
#         if not attr2_values.empty:
#             corrected_value = attr2_values.mode()[0]
#         else:
#             corrected_value = error['dirty_value']
        
#         if corrected_value != error['ground_truth']:
#             print(f"expect {error['ground_truth']}, get {corrected_value}")
#             flag = False
#             break
#     if flag == False:
#         print(f'{column} failed')
#         return flag
#     for idx in detection.index[detection[column] == 1]:
#         valid_rows = (dirty_data[attr1] == dirty_data.at[idx, attr1]) & (detection[attr1] == 0) & (detection[attr2] == 0)
#         attr2_values = dirty_data.loc[valid_rows, attr2]
#         if not attr2_values.empty:
#             corrected_value = attr2_values.mode()[0]
#         else:
#             corrected_value = corrections.at[idx, column]
#         corrections.at[idx, column] = corrected_value
#     print(f'{column} fd passed')
#     return True


def format_row(row, header):
    s = '{' + ', '.join(f'"{col}": "{val}"' for col, val in zip(header, row)) + '}'
    return s


def format_row_2(value, key, detection_row):
    result = {key[i]: value[i] for i in range(len(value)) if detection_row[i] == 0}
    return json.dumps(result)


def repair_table():
    # ------------------------------------------------------------------------------------------------------------------------
    # Generating CoTs for representative errors
    # dataset_specific_examples generated by llms
    # dataset_specific_examples = {}
    # representative_error_info sampled from dirty_data_human_repaired
    # d[column][index], 1. dirty_tuple 2. dirty_value 3. ground_truth 4. error_analysis 5. error_type
    # representative_error_info = {}
    # time_cot_start = time.time()
    # retriever_build_time = 0
    # print('Generating CoTs for representative errors...')
    # with ThreadPoolExecutor(max_workers=20) as executor:
    #     futures = []
    for col_idx, column in enumerate(detection.columns):
        if detection[column].sum() > 0:
            # print(column)
            # print('Building retriever...')
            # retriever_start = time.time()
            retriever, indices, CoE = build_retriever(dirty_data,
                                                        detection,
                                                        embeddings_matrix,
                                                        column,
                                                        )
            # retriever_build_time += time.time() - retriever_start
            # print('Building completed')
            retriever_dict[column] = retriever
            indices_dict[column] = indices
            CoE_dict[column] = CoE
                # detection_filtered = detection.iloc[:, indices]
                # future = executor.submit(get_auto_cot,
                #                          representative_tuples_list,
                #                          column,
                #                          retriever,
                #                          CoE,
                #                          indices,
                #                          detection_filtered,
                #                          representative_error_info,
                #                          dataset_specific_examples
                #                          )
        #         futures.append(future)
        # for future in as_completed(futures):
        #     future.result()
    # with open(os.path.join(output_path, 'dataset_specific_examples.txt'), 'w', encoding='utf-8') as f_output:
    #     for column, few_shot_dataset_specific_str in dataset_specific_examples.items():
    #         f_output.write(f"Column: {column}\n")
    #         f_output.write(f"Examples:\n{few_shot_dataset_specific_str}\n")
    #         f_output.write("\n" + "="*50 + "\n\n")
    # print('CoT generation completed') 
    # time_cot_end = time.time()
    # print('time cost: ', time_cot_end - time_cot_start)
    # print(f"retriever build time: {retriever_build_time}")
    # rep_clean_data = sel_clean(human_repair_num)
    # Merge the sampled representative errors and correct data
    # rep_data_info = copy.deepcopy(representative_error_info)
    # if not rep_clean_data.empty:
    #     clean_data_integration(rep_clean_data, rep_data_info)
    # ------------------------------------------------------------------------------------------------------------------------
    # Generate correction rules (including python code for string transformation and functional dependencies)
    # print('Start code generation')
    # time_code_generation_start = time.time()
    # codes = {}
    # fds = {}
    # train_data = {}
    # val_data = {}
    # with ThreadPoolExecutor(max_workers=20) as executor:
    #     futures = []
    #     for column in detection.columns:
    #         if detection[column].sum() > 0:
    #             formatting_issue = True
    #             for error in representative_error_info[column].values():
    #                 if error['error_type'] != 'clean':
    #                     if error['error_type'] != 'Formatting Issue':
    #                         formatting_issue=False
    #                         break
    #             train_data[column], val_data[column] = train_val_split(rep_data_info[column])
    #             if formatting_issue:
    #                 future = executor.submit(code_generation,
    #                                         train_data[column],
    #                                         column,
    #                                         codes
    #                                         )
    #             else:
    #                 future = executor.submit(FD_generation,
    #                                         train_data[column],
    #                                         column,
    #                                         fds
    #                                         )
    #             futures.append(future)
    #     for future in as_completed(futures):
    #         future.result()
    # with open(os.path.join(output_path, 'codes.txt'), 'w', encoding='utf-8') as f_output:
    #     for column, code in codes.items():
    #         f_output.write(f"Column: {column}\n")
    #         f_output.write(f"Code:\n{code}\n")
    #         f_output.write("\n" + "="*50 + "\n\n")
    # with open(os.path.join(output_path, 'fds.txt'), 'w', encoding='utf-8') as f_output:
    #     for column, fd in fds.items():
    #         f_output.write(f"Column: {column}\n")
    #         f_output.write(f"Fd:\n{json.dumps(fd)}\n")
    #         f_output.write("\n" + "="*50 + "\n\n")
    # time_code_generation_end = time.time()
    # print('Code generation completed')
    # print('time cost: ', time_code_generation_end - time_code_generation_start)
    # ------------------------------------------------------------------------------------------------------------------------
    # Evaluate and execute the generated code
    # print('Start code evaluation and execution')
    # time_code_evaluation_start = time.time()
    # for column in codes.keys():
    #     code_evaluation_execution(codes[column],rep_data_info[column],column)
    # for column in fds.keys():
    #     fd_evaluation_execution(fds[column],rep_data_info[column],column)
    # time_code_evaluation_end = time.time()
    # print('Code evaluation and execution completed')
    # print('time cost: ', time_code_evaluation_end - time_code_evaluation_start)
    # ------------------------------------------------------------------------------------------------------------------------
    # In the next step of repairing, only consider the elements that are not repaired by the code and functional dependencies
    # Compare corrections and dirty_data_human_repaired, find the elements in dirty_data_human_repaired that are repaired by corrections, then update detection_human_repaired, set these elements to 0
    # detection_human_repaired_copy = detection_human_repaired.copy()
    # mask = dirty_data_human_repaired != corrections
    # detection_human_repaired_copy[mask] = 0
    # # ------------------------------------------------------------------------------------------------------------------------
    # # Update the retriever
    # print('Start updating retriever')
    # time_update_retriever_start = time.time()
    # for col_idx, column in enumerate(detection_human_repaired.columns):
    #     if detection_human_repaired[column].sum() > 0:
    #         update_retriever(column)
    # time_update_retriever_end = time.time()
    # print('Updating retriever completed')
    # print('time cost: ', time_update_retriever_end - time_update_retriever_start)
    # ------------------------------------------------------------------------------------------------------------------------
    # Retrieve relevant tuples for each error
    print('Start retrieving')
    if dataset_name != 'imdb':
        time_retrieving_start = time.time()
        total_time = 0
        retriever_time = 0
        dict_creation_time = 0
        sort_time = 0
        for col_idx, column in enumerate(dirty_data.columns):
            if detection[column].sum() > 0:
                retriever = retriever_dict[column]
                indices = indices_dict[column]
                CoE = CoE_dict[column]
                temp = detection.iloc[:, indices]
                for row_idx in range(len(detection)):
                    # Only retrieve elements that are not repaired by functional dependencies, but the sorting needs to be based on the original Detection
                    if detection.at[row_idx, column] == 1:
                        # Convert to NumPy array in the outer loop
                        column_detection = detection[column].values
                        column_indices = indices_dict[column]
                        column_sums = detection.iloc[:, column_indices].values.sum(axis=1)
                        start_time = time.time()
                        relevant_clean_tuples = ''
                        embeddings_row = embeddings_matrix[row_idx]
                        embeddings_row_filtered = embeddings_row[indices]
                        for i in range(len(embeddings_row_filtered)):
                            if temp.iloc[row_idx, i] == 1:
                                embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
                        embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
                        embeddings_row_united = embeddings_row_filtered.flatten()
                        retriever_start = time.time()
                        relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
                                                                                        k=30)
                        retriever_time += time.time() - retriever_start
                        dict_start = time.time()
                        # Use NumPy array in the inner loop
                        relevant_rows_dict_list = [
                            {
                                'page_content': row[0].page_content,
                                'index': idx,
                                'score': round(row[1], 2),
                                'target_column': column_detection[idx],
                                'sum': column_sums[idx]
                            }
                            for row in relevant_rows
                            for idx in [row[0].metadata['index']]
                        ]
                        dict_creation_time += time.time() - dict_start
                        sort_start = time.time()
                        sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
                                                                    'sum')
                        sort_time += time.time() - sort_start
                        for row in sorted_relevant_rows_dict_list[:3]:
                            relevant_clean_tuples += row['page_content'] + '\n'
                        if column not in retrieved_tuples:
                            retrieved_tuples[column] = {}
                        retrieved_tuples[column][row_idx] = relevant_clean_tuples
                        # total_time += time.time() - start_time
        # print(f"Total time: {total_time:.2f} seconds")
    # print(f"Retriever time: {retriever_time:.2f} seconds ({retriever_time / total_time * 100:.2f}%)")
    # print(f"Dict creation time: {dict_creation_time:.2f} seconds ({dict_creation_time / total_time * 100:.2f}%)")
    # print(f"Sort time: {sort_time:.2f} seconds ({sort_time / total_time * 100:.2f}%)")
    # print(
    #     f"Other operations: {(total_time - retriever_time - dict_creation_time - sort_time):.2f} seconds ({(total_time - retriever_time - dict_creation_time - sort_time) / total_time * 100:.2f}%)")
    print('Retrieval completed')
    # time_retrieving_end = time.time()
    # print('time cost: ', time_retrieving_end - time_retrieving_start)
    # ------------------------------------------------------------------------------------------------------------------------
    # Repair the remaining errors using LLMs
    print('Repairing the remaining errors using LLMs...')
    time_repairing_start = time.time()
    with ThreadPoolExecutor(max_workers=50) as executor:
        futures = []
        for col_idx, column in enumerate(detection.columns):
            # If there are errors in the column
            if detection[column].sum() > 0:
                # First use a loop to form the Input, then pass the Input into the prompt
                # Iterate through each element in the column
                prompt = ChatPromptTemplate(
                    messages=[
                        SystemMessagePromptTemplate.from_template(sys_EC),
                        HumanMessagePromptTemplate.from_template(human_EC),
                    ],
                    partial_variables={
                        "general_examples": general_examples_EC_str,
                        "dataset_specific_examples": ' ',
                        "format_instructions": parser.get_format_instructions()
                    }
                )
                chain = (
                        prompt
                        | LLM4EC
                        | parser
                )
                for row_idx in range(len(detection)):
                    if detection.at[row_idx, column] == 1:
                        # dirty tuple
                            dirty_tuple = dirty_data.iloc[row_idx]
                            dirty_value = dirty_data.at[row_idx, column]
                            # repair the dirty value
                            future = executor.submit(repair_value,
                                                     dirty_tuple,
                                                     column,
                                                     dirty_value,
                                                     row_idx,
                                                     col_idx,
                                                     chain,
                                                     )
                            futures.append(future)
        for future in as_completed(futures):
            future.result()
    time_repairing_end = time.time()
    print('Repairing completed')
    print('time cost: ', time_repairing_end - time_repairing_start)
    # ------------------------------------------------------------------------------------------------------------------------


def repair_value(dirty_tuple, column, dirty_value, index_row, index_col, chain):
    # Invoke the LLM to repair a dirty value, return the repaired value and the chain-of-thought for it
    
    # Filter irrelevant columns from the dirty tuple
    filtered_tuple = dirty_tuple.iloc[indices_dict[column]]
    filtered_header = [header[i] for i in indices_dict[column]]
    # Get the string form of the filtered dirty tuple
    dirty_tuple_filtered_str = format_row(filtered_tuple, filtered_header)
    # Get the string form of the dirty value
    dirty_value_str = '{' + f'"{column}": "{dirty_value}"' + '}'
    # Get the JSON form of the dirty tuple
    dirty_tuple_json = dirty_tuple.to_dict()
    # Initialize the correction and result
    correction = dirty_value
    result_json = dirty_value_str
    # Get the relevant clean tuples
    if dataset_name != 'imdb':
        relevant_clean_tuples = retrieved_tuples[column][index_row]
    else:
        relevant_clean_tuples = ' '
    # LLM invocations can be unstable, so we try multiple times
    try_num = 0
    while True:
        try:
            # Invoke the LLM to repair the dirty value
            repair_result = chain.invoke({'Dirty_Tuple': dirty_tuple_filtered_str,
                                          'Erroneous_value': dirty_value_str,
                                          'Relevant_clean_tuples': relevant_clean_tuples,
                                          })
            result_json = repair_result
            # If the column is not in the correction, set it to 'null'
            if result_json['correction'].get(column) is None:
                result_json['correction'][column] = 'null'
            # Update the correction
            correction = result_json['correction'][column]
            break
        except Exception as e:
            print('ChatModel request error', e)
            try_num += 1
            # If the number of attempts exceeds 3, break
            if try_num >= 3:
                break
            continue
    # Update the correction
    corrections.iloc[index_row, index_col] = str(correction)
    # Log the repair process
    log = {'Index': dirty_tuple_json['index'],
           'Dirty_tuple': format_row(dirty_tuple, header),
           'Dirty_value': dirty_value_str,
           'Relevant_clean_tuples': relevant_clean_tuples,
           'Correction': str(result_json)
           }
    logs.append(log)


def cmp_mark(df_A, df_B):
    # Highlight all inconsistent elements
    df_A.fillna('null', inplace=True)
    df_B.fillna('null', inplace=True)
    # Find different elements
    difference = df_A.ne(df_B)
    # Calculate the total number of inconsistent elements
    diff_count = difference.sum().sum()
    # Calculate the percentage of inconsistent elements
    total_elements = difference.size
    diff_percent = diff_count / total_elements * 100
    # Create a new Excel file
    wb = Workbook()
    ws = wb.active
    # Fill the data of the B DataFrame into the worksheet
    for r in dataframe_to_rows(df_B, index=False, header=True):
        ws.append(r)
    # Iterate through the difference DataFrame, add highlighting to the cells in B that are different from A
    for col in range(difference.shape[1]):
        # Note: openpyxl starts counting from 1, so we need to add 2 (1 for the first row, and 1 for skipping the title row)
        for row in range(difference.shape[0]):
            if difference.iloc[row, col]:
                cell = ws.cell(row=row + 2, column=col + 1)
                # Set the fill color to yellow
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Print the number and percentage of inconsistent elements
    print(f"Number of inconsistent elements: {diff_count}, Percentage: {diff_percent:.2f}%")
    # Save as an xlsx file
    filename_with_extension = os.path.basename(dirty_data_path)
    filename_without_extension = os.path.splitext(filename_with_extension)[0]
    save_path = os.path.join(output_path, f'{filename_without_extension}-corrected-marked_{MODEL_NAME_EC}.xlsx')
    wb.save(save_path)


def harmonic_mean(a, b):
    if a == 0 or b == 0:
        return 0
    return 2 / (1 / a + 1 / b)


def calc_p_r_f(clean_data, dirty_data, corrected_data):
    # Calculate P, R, F
    # When corrected_data and dirty_data are inconsistent, it indicates that the model has repaired the element
    mask_bc = corrected_data != dirty_data
    # When corrected_data and clean_data are consistent, it indicates that the model has repaired the element successfully
    mask_ac = clean_data == corrected_data
    corrected_num = mask_bc.sum().sum()
    final_mask1 = mask_ac & mask_bc
    right_corrected_num1 = final_mask1.sum().sum()
    Precision = right_corrected_num1 / corrected_num
    # When clean_data and dirty_data are inconsistent, it indicates that the element is incorrect
    mask_ab = clean_data != dirty_data
    dirty_num = mask_ab.sum().sum()
    # When corrected_data and clean_data are consistent, it indicates that the error element has been successfully repaired
    final_mask2 = mask_ab & mask_ac
    right_corrected_num2 = final_mask2.sum().sum()
    Recall = right_corrected_num2 / dirty_num
    F1 = harmonic_mean(Precision, Recall)
    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        print(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}')
        f_output.write(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}\n')


def save_logs(logs):
    with open(os.path.join(output_path, 'output.txt'), 'w', encoding='utf-8') as f_output:
        for log in logs:
            f_output.write(f"Dirty_tuple:\n{log['Dirty_tuple']}\n")
            f_output.write(f"Dirty_value:\n{log['Dirty_value']}\n")
            f_output.write(f"Relevant_clean_tuples:\n{log['Relevant_clean_tuples']}\n")
            f_output.write(f"Correction:\n{log['Correction']}\n")


if __name__ == "__main__":
    # param-labeling_budget
    human_repair_num = 0
    # param-dataset_name
    dataset_name = 'hospital'
    # param-csv_file_path
    clean_data_path = f'datasets/{dataset_name}/{dataset_name}_clean.csv'
    dirty_data_path = f'datasets/{dataset_name}/{dataset_name}_dirty.csv'
    detection_path = f'datasets/{dataset_name}/{dataset_name}_dirty_error_detection.csv'
    output_path = get_folder_name(f'runs_{dataset_name}')
    # param-prompts_path
    system_message_EC_path = 'prompt_templates/SystemMessage_EC.txt'
    human_message_EC_path = 'prompt_templates/HumanMessage_EC.txt'
    general_examples_EC_path = 'prompt_templates/general_examples_EC.txt'
    # system_message_CoT_path = 'prompt_templates/SystemMessage_CoT_with_error_type.txt'
    # general_examples_CoT_path = 'prompt_templates/examples_CoT_with_error_type.txt'
    # human_message_CoT_large_path = 'prompt_templates/HumanMessage_CoT_large.txt'
    # human_message_CoT_small_path = 'prompt_templates/HumanMessage_CoT_small.txt'
    # system_message_code_generation_path = 'prompt_templates/SystemMessage_code_generation.txt'
    # human_message_code_generation_path = 'prompt_templates/HumanMessage_code_generation.txt'
    # system_message_fd_generation_path = 'prompt_templates/SystemMessage_fd_generation.txt'
    # human_message_fd_generation_path = 'prompt_templates/HumanMessage_fd_generation.txt'
    # param-llm4EC
    MODEL_NAME_EC = 'gpt-3.5-turbo-0125'
    OPENAI_API_BASE_EC = ''
    OPENAI_API_KEY_EC = ''
    TEMPERATURE_EC = 0
    # MODEL_NAME_CoT = 'gpt-4o-2024-11-20'
    # OPENAI_API_BASE_CoT = ''
    # OPENAI_API_KEY_CoT = ''
    # TEMPERATURE_CoT = 0
    # param-llm4Code
    # MODEL_NAME_Code = 'gpt-4o-2024-11-20'
    # OPENAI_API_BASE_Code = ''
    # OPENAI_API_KEY_Code = ''
    # TEMPERATURE_Code = 0
    # param-llm4FD
    # MODEL_NAME_FD = 'gpt-4o-2024-11-20'
    # OPENAI_API_BASE_FD = ''
    # OPENAI_API_KEY_FD = ''
    # TEMPERATURE_FD = 0
    # param-embedding_model
    EMBEDDING_MODEL_PATH = r"xxxx\xxxx\all-MiniLM-L6-v2"
    # ------------------------------------------------------------------------------------------------------------------------
    # load data
    clean_data = pd.read_csv(clean_data_path, dtype=str, encoding='utf-8')
    clean_data.fillna('null', inplace=True)
    dirty_data = pd.read_csv(dirty_data_path, dtype=str, encoding='utf-8')
    dirty_data.fillna('null', inplace=True)
    header = dirty_data.columns.tolist()
    row_count, column_count = dirty_data.shape  
    detection = pd.read_csv(detection_path)
    print('row_count:', row_count)
    print('column_count:', column_count)
    print('dirty value num:', detection.sum().sum())
    print('dirty value rate:', detection.sum().sum() / row_count / column_count)
    # ------------------------------------------------------------------------------------------------------------------------
    # load prompts and predefined examples
    general_examples_EC_str = load_examples(general_examples_EC_path)
    # general_examples_CoT_str = load_examples(general_examples_CoT_path)
    parser = JsonOutputParser(pydantic_object=Output)
    sys_EC, human_EC = load_prompts(system_message_EC_path, human_message_EC_path)
    # sys_CoT, human_CoT_large, human_CoT_small = load_prompts(
    #     system_message_CoT_path,
    #     human_message_CoT_large_path,
    #     human_message_CoT_small_path
    # )
    # sys_code_generation, human_code_generation = load_prompts(system_message_code_generation_path, human_message_code_generation_path)
    # sys_fd_generation, human_fd_generation = load_prompts(system_message_fd_generation_path, human_message_fd_generation_path)
    # ------------------------------------------------------------------------------------------------------------------------  
    # Initialize the ChatOpenAI instance for standard language model with specified configurations
    LLM4EC = initialize_llm(MODEL_NAME_EC, OPENAI_API_BASE_EC, OPENAI_API_KEY_EC, TEMPERATURE_EC)
    # LLM4CoT = initialize_llm(MODEL_NAME_CoT, OPENAI_API_BASE_CoT, OPENAI_API_KEY_CoT, TEMPERATURE_CoT)
    # LLM4Code = initialize_llm(MODEL_NAME_Code, OPENAI_API_BASE_Code, OPENAI_API_KEY_Code, TEMPERATURE_Code)
    # LLM4FD = initialize_llm(MODEL_NAME_FD, OPENAI_API_BASE_FD, OPENAI_API_KEY_FD, TEMPERATURE_FD)
    # # ------------------------------------------------------------------------------------------------------------------------
    # start repair
    logs = []
    corrections = dirty_data.copy()
    start_time = time.time()
    # embedding
    print('embedding...')
    emb_start_time = time.time()
    embeddingModel = myEmbeddings(EMBEDDING_MODEL_PATH)
    embeddings_matrix, embedding_dimension = embed_data(dirty_data, embeddingModel)
    # embeddings_matrix = np.load('embeddings_matrix_imdb.npy')
    embedding_dimension = embeddings_matrix.shape[2]
    # print('calc stasts..')
    # stasts_features = calc_stasts(dirty_data)
    # print('done')
    emb_end_time = time.time()
    print('embedding done')
    print(f"embedding time cost {emb_end_time-emb_start_time}")
    # ------------------------------------------------------------------------------------------------------------------------
    # select representative tuples
    # print('selecting representative tuples...')
    # select_start_time = time.time()
    # representative_tuples_list = select_representative_tuples_old(embeddings_matrix, detection, human_repair_num)
    # print(representative_tuples_list)
    # Print the selected representative tuples' "ounces" and "abv" columns
    # print("\nSelected Representative Tuples:")
    # if "ounces" in dirty_data.columns and "abv" in dirty_data.columns:
    #     representative_data = dirty_data.iloc[representative_tuples_list][["ounces", "abv"]]
    #     print(representative_data)
    # else:
    #     if "ounces" not in dirty_data.columns:
    #         print("Column 'ounces' not found in the dataset")
    #     if "abv" not in dirty_data.columns:
    #         print("Column 'abv' not found in the dataset")
    #     # Print available columns if the requested ones don't exist
    #     print(f"Available columns: {dirty_data.columns.tolist()}")
    # select_end_time = time.time()
    # print('select done')
    # print(f"select time cost {select_end_time-select_start_time}")
    # simulation of human repair
    # dirty_data_human_repaired = dirty_data.copy()
    # dirty_data_human_repaired.iloc[representative_tuples_list] = clean_data.iloc[representative_tuples_list]
    # detection_human_repaired = detection.copy()
    # detection_human_repaired.iloc[representative_tuples_list] = 0
    # corrections.iloc[representative_tuples_list] = clean_data.iloc[representative_tuples_list]
    # dirty_data_only_repaired = clean_data.iloc[representative_tuples_list]
    # elements_list_only_repaired = dirty_data_only_repaired.values.flatten().tolist()
    # embeddings_only_repaired = embeddingModel.embed_documents(elements_list_only_repaired)
    # embeddings_matrix_only_repaired = embeddings_only_repaired.reshape(human_repair_num, column_count, embedding_dimension)
    # ------------------------------------------------------------------------------------------------------------------------
    retriever_dict = {}
    indices_dict = {}
    CoE_dict = {}
    retrieved_tuples = {}
    repair_table()   
    end_time = time.time()
    logs = sorted(logs, key=lambda x: int(x['Index']))
    save_logs(logs)
    corrections.to_csv(os.path.join(output_path, 'corrections.csv'), encoding='utf-8',index=False)
    cmp_mark(clean_data, corrections)
    calc_p_r_f(clean_data, dirty_data, corrections)
    execution_time = end_time - start_time
    print(f"execution time: {execution_time} seconds")
    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        f_output.write(f"Time consumption: {execution_time}\n")