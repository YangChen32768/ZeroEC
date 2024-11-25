########################################
# ZeroEC: The training-free and explainable error correction system powered by large language models
# Chen Yang
# yc32768@gmail.com
# September 2024
# Software College
# Zhejiang University
# All Rights Reserved
########################################

########################################
from cmath import sqrt
import threading

from langchain.prompts import (
    PromptTemplate,
    ChatPromptTemplate,
    HumanMessagePromptTemplate,
    SystemMessagePromptTemplate,
    MessagesPlaceholder,
    FewShotChatMessagePromptTemplate,
)
from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, BaseMessage
from langchain.memory import ConversationBufferMemory, ConversationBufferWindowMemory
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain_core.output_parsers import JsonOutputParser
from langchain_community.document_loaders.csv_loader import CSVLoader
from langchain_community.vectorstores import FAISS
from langchain_community.retrievers import BM25Retriever
from langchain.retrievers import EnsembleRetriever
from langchain_core.runnables import RunnableLambda, RunnablePassthrough
from operator import itemgetter
import ast
from networkx import goldberg_radzik
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import time
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import concurrent.futures
import pprint
import numpy as np
from langchain_community.vectorstores import FAISS
import pandas as pd
from sklearn.metrics import mutual_info_score
from scipy.stats import entropy
from sentence_transformers import SentenceTransformer
from langchain_core.embeddings import Embeddings
from typing import List
import csv
from sklearn.cluster import KMeans,MiniBatchKMeans
from fast_sentence_transformers import FastSentenceTransformer
from langchain_community.vectorstores.utils import DistanceStrategy
import copy
from collections import defaultdict
import random
from collections import Counter
from dashscope import get_tokenizer

########################################

class Output(BaseModel):
    # The output parser
    chain_of_thought_for_correction: str = Field(...,
                                                 description="The chain_of_thought that led to the proposed correction")
    correction: dict = Field(...,
                             description="the most probable correction for the dirty value")


def form_examples(examps) -> str:
    # 接受一个examples数据，返回用于prompt的字符串
    # examps是一个数组，其中每个元素都是包含了'input'和'output'的字典
    few_shot_examps_str = ''
    for examp in examps:
        examp_str = 'human: ' + \
                    json.dumps(examp['input']) + \
                    '\n' + \
                    'ai: ' + \
                    json.dumps(examp['output']) + \
                    '\n'
        few_shot_examps_str = few_shot_examps_str + examp_str
    return few_shot_examps_str


def get_folder_name(base_path):
    # 获取本次运行时的保存路径
    if not os.path.exists(base_path):
        print(f'base_path:{base_path} does no exist!\nTrying to make a new one.')
        os.makedirs(base_path)
    existing_folders = [d for d in os.listdir(base_path) if
                        os.path.isdir(os.path.join(base_path, d)) and d.startswith('run-')]
    max_run = 0
    for folder in existing_folders:
        run_number = int(folder.split('-')[1])
        if run_number > max_run:
            max_run = run_number
    next_run_folder = f'run-{max_run + 1}'
    next_run_path = os.path.join(base_path, next_run_folder)
    # 创建新的运行文件夹
    os.makedirs(next_run_path)
    return next_run_path


def select_repair_candidates(embeddings_matrix: np.ndarray, detection: pd.DataFrame, num_clusters: int) -> list:
    # Select rows to repair based on detection
    mask_rows = detection.sum(axis=1) > 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    filtered_detection = detection[mask_rows]

    # Store original indices
    original_indices = np.where(mask_rows)[0]

    # Reshape embeddings
    m, n, l = filtered_embeddings.shape
    mask = filtered_detection.astype(bool)
    masked_embeddings = filtered_embeddings * mask.values[..., np.newaxis]
    reshaped_embeddings = masked_embeddings.reshape((m, n * l))

    # Perform clustering
    kmeans = MiniBatchKMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)

    # Greedy selection of candidates from each cluster
    selected_indices = []
    covered_columns = set()

    for i in range(num_clusters):
        cluster_mask = clusters == i
        if np.any(cluster_mask):
            cluster_detection = filtered_detection.iloc[cluster_mask]
            cluster_original_indices = original_indices[cluster_mask]

            # Greedy selection within cluster
            best_index = None
            best_new_coverage = -1

            for idx, orig_idx in zip(cluster_detection.index, cluster_original_indices):
                row = cluster_detection.loc[idx]
                new_columns = set(row[row == 1].index) - covered_columns
                if len(new_columns) > best_new_coverage:
                    best_new_coverage = len(new_columns)
                    best_index = idx
                    best_orig_index = orig_idx

            if best_index is not None:
                selected_indices.append(best_orig_index)
                covered_columns.update(
                    set(cluster_detection.loc[best_index][cluster_detection.loc[best_index] == 1].index))

    return selected_indices


def build_retriever_3(dirty_data, detection, vectors_matrix, target_column):
    # 记录代码开始执行的时间
    start_time = time.time()

    # 1. texts 转换
    texts = dirty_data.values.tolist()
    print(f"Step 1: Converting dirty data to texts - Time Taken: {time.time() - start_time:.4f}s")

    # 2. 计算互信息
    step_start_time = time.time()
    normMI = calc_mi_2(dirty_data, target_column) # 10s
    normMI = np.round(normMI, decimals=1)
    print(f"Step 2: Calculating Mutual Information (MI) - Time Taken: {time.time() - step_start_time:.4f}s")
    
    print(normMI)
    normMI = np.array(normMI)
    threshold = 0.5

    # 3. 筛选相关列
    step_start_time = time.time()
    indices = np.where(normMI >= threshold)[0]
    CoE = normMI[indices]
    embeddings_matrix_col_filtered = vectors_matrix[:, indices, :]
    texts_col_filtered = [[row[col] for col in indices] for row in texts]
    detection_col_filtered = detection.iloc[:, indices]
    print(f"Step 3: Filtering columns based on MI - Time Taken: {time.time() - step_start_time:.4f}s")

    # 4. 合并 texts
    step_start_time = time.time()
    formatted_rows = []
    header_col_filtered = detection_col_filtered.columns
    for text_row, detection_row in zip(texts_col_filtered, detection_col_filtered.values.tolist()):
        formatted_rows.append(format_row_2(text_row, header_col_filtered, detection_row))
    print(f"Step 4: Merging texts with formatted detection - Time Taken: {time.time() - step_start_time:.4f}s")

    # 5. 重塑 embeddings_matrix
    step_start_time = time.time()
    embeddings_matrix_col_row_filtered_reshaped = embeddings_matrix_col_filtered.reshape(
        embeddings_matrix_col_filtered.shape[0], -1)
    print(f"Step 5: Reshaping embeddings matrix - Time Taken: {time.time() - step_start_time:.4f}s")

    # 6. 生成配对数据
    step_start_time = time.time()
    # 使用生成器表达式来避免一次性占用大量内存
    paired_data_generator = ((text_row, vector) for text_row, vector in zip(formatted_rows, embeddings_matrix_col_row_filtered_reshaped))

    # 如果后续需要将其转换为列表，可以按需进行
    paired_data = list(paired_data_generator)

    print(f"Step 6: Creating paired data - Time Taken: {time.time() - step_start_time:.4f}s")

    # 7. 生成 IDs
    step_start_time = time.time()
    ids = [str(i) for i in range(len(dirty_data_human_repaired))]
    print(f"Step 7: Generating IDs - Time Taken: {time.time() - step_start_time:.4f}s")

    # 8. 生成 meta_data
    step_start_time = time.time()
    meta_data = [{'index': i} for i in range(len(dirty_data_human_repaired))]
    print(f"Step 8: Generating meta data - Time Taken: {time.time() - step_start_time:.4f}s")

    # 9. 创建 FAISS 索引
    step_start_time = time.time()
    db = FAISS.from_embeddings(text_embeddings=paired_data,
                               embedding=embeddingModel,
                               metadatas=meta_data,
                               ids=ids,
                               distance_strategy=DistanceStrategy.DOT_PRODUCT)
    retriever = db
    print(f"Step 9: Creating FAISS index - Time Taken: {time.time() - step_start_time:.4f}s")

    # 输出整体执行时间
    print(f"Total Time Taken: {time.time() - start_time:.4f}s")

    return retriever, indices, CoE


def update_retriever(column):
    # 利用用户修复的结果更新检索器
    ids = [str(i) for i in repair_list]
    retriever_dict[column].delete(ids)
    embeddings_matrix_only_repaired_col_filtered = embeddings_matrix_only_repaired[:, indices_dict[column], :]
    texts = dirty_data_only_repaired.values.tolist()
    texts_col_filtered = [[row[col] for col in indices_dict[column]] for row in texts]
    header_col_filtered = [header[i] for i in indices_dict[column]]
    formatted_rows = []
    for text_row in texts_col_filtered:
        formatted_rows.append(format_row(text_row, header_col_filtered))
    meta_data = [{'index': i} for i in repair_list]
    paired_data = [(text, vector.flatten().tolist()) for text, vector in
                   zip(formatted_rows, embeddings_matrix_only_repaired_col_filtered)]
    retriever_dict[column].add_embeddings(text_embeddings=paired_data,
                                          metadatas=meta_data,
                                          ids=ids
                                          )


def sort_dicts(dict_list, key1, key2, key3):
    return sorted(dict_list, key=lambda x: (x.get(key1, float('inf')),
                                            x.get(key2, float('inf')),
                                            x.get(key3, float('inf'))))


def get_auto_cot(repair_list, column, retriever, CoE, indices, detection_filtered, rep_error_info: dict, sp_examps):
    rep_error_info[column] = {}
    human_message_small_template = HumanMessagePromptTemplate.from_template(human_auto_cot_small)
    human_input = '['
    relevant_rows_list = []
    dirty_tuples = []
    dirty_tuples_list = []
    dirty_values = []
    clean_values = []
    filtered_header = [header[i] for i in indices]
    column_detection = detection_human_repaired[column].values
    column_indices = indices_dict[column]
    column_sums = detection_human_repaired.iloc[:, column_indices].values.sum(axis=1)
    for row_idx in repair_list:
        if detection.loc[row_idx, column] == 1:
            dirty_value = dirty_data.loc[row_idx, column]
            dirty_values.append(dirty_value)
            clean_value = clean_data.loc[row_idx, column]
            clean_values.append(clean_value)
            dirty_row = dirty_data.iloc[:, indices].loc[row_idx].tolist()
            dirty_tuples_list.append(dirty_row)
            dirty_tuples.append(dirty_data.iloc[:, indices].loc[row_idx])
            relevant_clean_tuples = ''
            embeddings_row = embeddings_matrix[row_idx]
            embeddings_row_filtered = embeddings_row[indices]
            for i in range(len(embeddings_row_filtered)):
                if detection_filtered.iloc[row_idx, i] == 1:
                    embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
            embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
            embeddings_row_united = embeddings_row_filtered.flatten()
            relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
                                                                             k=30)
            relevant_rows_dict_list = [
                {
                    'page_content': row[0].page_content,
                    'index': idx,
                    'score': round(row[1], 2),
                    'target_column': column_detection[idx],
                    'sum': column_sums[idx]
                }
                for row in relevant_rows
                for idx in [row[0].metadata['index']]
            ]

            sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
                                                        'sum')
            for row in sorted_relevant_rows_dict_list[:3]:
                relevant_clean_tuples += row['page_content'] + '\n'
            relevant_rows_list.append(relevant_clean_tuples)
            human_input += '{' + human_message_small_template.format(
                Dirty_Tuple = format_row(dirty_row, filtered_header),
                Erroneous_value = '{' + f'"{column}": "{dirty_value}"' + '}',
                Relevant_clean_tuples = relevant_clean_tuples,
                Correction = clean_value
            ).content + '},'
    human_input += ']'
    prompt_auto_cot = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(sys_auto_cot),
        HumanMessagePromptTemplate.from_template(human_auto_cot_large),
    ],
        partial_variables={
            "examples": examples_auto_cot_str,
        }
    )
    chain_auto_cot = (prompt_auto_cot | llm_auto_cot)
    while True: 
        try:
            repair_result = chain_auto_cot.invoke(
                {"human_input": human_input}).content
            repair_result = repair_result.replace('\n', '')
            repair_result = re.sub(r'```json', '', repair_result)
            repair_result = re.sub(r'```', '', repair_result)
            repair_result_list = ast.literal_eval(repair_result)
            break
        except Exception as e:
            print('auto-cot failed', e)
    
    # 构建examples
    specific_examples_llm = []
    for idx, result in enumerate(repair_result_list):
        dirty_tuple = dirty_tuples_list[idx]
        json_string = """
                            {
                                "input": {
                                    "Dirty Data Tuple": null,
                                    "Erroneous Value": null,
                                    "Relevant Clean Tuples": null
                                },
                                "output": {
                                    "chain_of_thought_for_correction": null,
                                    "correction": null
                                }
                            }
                      """
        d = json.loads(json_string)
        d['input']['Dirty Data Tuple'] = format_row(dirty_tuple, filtered_header)
        d['input'][
            'Erroneous Value'] = '{' + f'"{column}": "{dirty_values[idx]}"' + '}'
        d['input']['Relevant Clean Tuples'] = relevant_rows_list[idx],
        d['output']['chain_of_thought_for_correction'] = result['chain_of_thought_for_correction']
        d['output']['correction'] = '{"' + column + '\": ' + '"' + clean_values[idx] + '"}'
        rep_error_info[column][dirty_tuples[idx].name] = {}
        rep_error_info[column][dirty_tuples[idx].name]['dirty_tuple'] = dirty_tuples[idx].to_dict()
        rep_error_info[column][dirty_tuples[idx].name]['dirty_value'] = dirty_values[idx]
        rep_error_info[column][dirty_tuples[idx].name]['ground_truth'] = clean_values[idx]
        rep_error_info[column][dirty_tuples[idx].name]['error_analysis'] = result['chain_of_thought_for_correction']
        rep_error_info[column][dirty_tuples[idx].name]['error_type'] = result['error_type']
        specific_examples_llm.append(d)
    few_shot_specific_str = form_examples(specific_examples_llm)
    sp_examps[column] = few_shot_specific_str


# def data_augmentation(selected_tuples: pd.DataFrame, selected_tuples_detection: pd.DataFrame,selected_tuples_clean: pd.DataFrame, error_analysis: dict, column: str,data_augmentation_result_all: dict):
    # 使用LLM对selected_tuples进行数据增强，返回增强后的数据
    # indices = indices_dict[column]
    # filtered_header = [header[i] for i in indices]
    # filtered_tuples = selected_tuples.iloc[:, indices]
    # filtered_tuples_detection = selected_tuples_detection.iloc[:, indices]
    # # filtered_tuples_str = [format_row(filtered_tuples.iloc[i], filtered_header) for i in range(len(filtered_tuples))]
    # error_analysis_column = error_analysis[column]
    # # 生成数据增强的prompt，包含：0. Instruction 1. Dirty tuple 2. Erroneous value 3. Error analysis
    # prompt_data_augmentation = ChatPromptTemplate(messages=[
    #     SystemMessagePromptTemplate.from_template(sys_data_augmentation),
    #     HumanMessagePromptTemplate.from_template(human_data_augmentation),
    # ])
    # chain_data_augmentation = (prompt_data_augmentation | llm_data_augmentation)
    # data_augmentation_result_column  = []
    # for i in range(len(filtered_tuples)):
    #     if filtered_tuples_detection[column].iloc[i] == 1:
    #         dirty_tuple_str = format_row(filtered_tuples.iloc[i], filtered_header)
    #         dirty_value = '{' + f'"{column}": "{filtered_tuples[column].iloc[i]}"' + '}'
    #         analysis = error_analysis_column[filtered_tuples.iloc[i].name]
    #         data_augmentation_result=(chain_data_augmentation.invoke({'Dirty_tuple': dirty_tuple_str,
    #                                                             'Dirty_value': dirty_value,
    #                                                             'Error_analysis': analysis,
    #                                                             }).content)
    #         data_augmentation_result = data_augmentation_result.replace('\n', '')
    #         data_augmentation_result = re.sub(r'```json', '', data_augmentation_result)
    #         data_augmentation_result = re.sub(r'```', '', data_augmentation_result)
    #         data_augmentation_result = re.sub(r'�','', data_augmentation_result)
    #         data_augmentation_result = ast.literal_eval(data_augmentation_result)
    #         original_error = {}
    #         original_error['generated_error'] = filtered_tuples.iloc[i].to_dict()
    #         original_error['error_analysis'] = analysis
    #         original_error['correct_value'] = {column: selected_tuples_clean[column].iloc[i]}
    #         data_augmentation_result_list = []
    #         data_augmentation_result_list.append(original_error)
    #         data_augmentation_result_list.append(data_augmentation_result)
    #         data_augmentation_result_column.append(data_augmentation_result_list)
    # data_augmentation_result_all[column] = data_augmentation_result_column
    # pass


def sel_clean(num_clusters:int):
    # Cluster the correct tuples and randomly select one from each cluster
    # Create a mask to select rows with zero errors
    mask_rows = detection.sum(axis=1) == 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    
    # Reshape embeddings for clustering
    m, n, l = filtered_embeddings.shape
    reshaped_embeddings = filtered_embeddings.reshape(m, n * l)
    if m == 0:
        return pd.DataFrame()
    if m < num_clusters:
        num_clusters = m
    # Perform KMeans clustering
    kmeans = MiniBatchKMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)

    # Get original indices of the selected clusters
    original_indices = np.where(mask_rows)[0]
    
    # Randomly select one sample from each cluster
    selected_indices = [
        original_indices[np.random.choice(np.where(clusters == cluster_id)[0])]
        for cluster_id in range(num_clusters)
        if np.any(clusters == cluster_id)
    ]
    
    return dirty_data.iloc[selected_indices]


def train_val_split(data:dict): 
    train_data = {}
    val_data = {}
    keys = list(data.keys())
    # 打乱字典的键
    random.shuffle(keys)
    split_ratio = 0.5
    # 计算切分索引
    split_index = int(len(data) * split_ratio)
    # 分割键列表
    train_keys = keys[:split_index]
    val_keys = keys[split_index:]
    # 划分数据集
    train_data = {key: data[key] for key in train_keys}
    val_data = {key: data[key] for key in val_keys}
    return train_data, val_data 


def clean_data_integration(clean_data:pd.DataFrame, rep_data_info:dict):
    # 对于各出错列，将采样所得的clean_data与dirty_data融合
    for column in dirty_data.columns:
         if detection[column].sum() > 0:
             for idx in range(len(clean_data)):
                rep_data_info[column][clean_data.iloc[idx].name] = {}
                rep_data_info[column][clean_data.iloc[idx].name]['dirty_tuple'] = clean_data.iloc[idx].to_dict()
                rep_data_info[column][clean_data.iloc[idx].name]['dirty_value'] = clean_data.iloc[idx][column]
                rep_data_info[column][clean_data.iloc[idx].name]['error_analysis'] = 'This is a clean value that does not need correction.'
                rep_data_info[column][clean_data.iloc[idx].name]['ground_truth'] = clean_data.iloc[idx][column]
                rep_data_info[column][clean_data.iloc[idx].name]['error_type'] = 'clean'
    pass


def code_generation(train_data:dict,column:str,codes:dict):
    # 若为formatting issue, 则让LLM生成代码
    # indices = indices_dict[column]
    # filtered_header = [header[i] for i in indices]
    examples = ''
    for error in train_data.values():
        # dirty_tuple_str = format_row(train_data[i]['generated_error'], filtered_header)
        dirty_value_str = error['dirty_value']
        analysis = error['error_analysis']
        clean_value_str = error['ground_truth']
        # examples += f"Dirty_tuple: {dirty_tuple_str}\n"
        examples += f"[Erroneous_value: {dirty_value_str}\n"
        examples += f"Error_analysis: {analysis}\n"
        examples += f"Correct_value: {clean_value_str}]\n"
    prompt_code_generation = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(sys_code_generation),
        HumanMessagePromptTemplate.from_template(human_code_generation),
        ])
    chain_code_generation = (prompt_code_generation | llm_code_generation)
    while True:
        try:
            raw_code = chain_code_generation.invoke({'examples': examples}).content
            extracted_code = re.findall(r'```python(.*?)```', raw_code, re.DOTALL)
            if extracted_code:
                codes[column] = extracted_code[0]
            break
        except Exception as e:
            print('code generation failed', e)


def correct():
    pass


def FD_generation(train_data:dict,column:str,FDs:dict):
    # 针对VAD,MISS,TYPO,寻找潜在的FD以完成修复
    indices = indices_dict[column]
    filtered_header = [header[i] for i in indices]
    examples = ''
    for error in train_data.values():
        dirty_tuple_str = json.dumps(error['dirty_tuple'])
        dirty_value_str = f"{column}: {error['dirty_value']}"
        analysis = error['error_analysis']
        correct_value = f"{column}: {error['ground_truth']}"
        examples += f"Dirty_tuple: {dirty_tuple_str}\n"
        examples += f"Erroneous_value: {dirty_value_str}, "
        examples += f"Error_analysis: {analysis}\n"
        examples += f"Correct_value: {correct_value}]\n"
    prompt_code_generation = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(sys_fd_generation),
        # HumanMessagePromptTemplate.from_template(human_fd_generation),
        ])
    chain_fd_generation = (prompt_code_generation | llm_code_generation)
    while True:
        try:
            fd_raw = chain_fd_generation.invoke({'examples': examples}).content
            fd_raw = re.sub(r'```json', '', fd_raw)
            fd_raw = re.sub(r'```', '', fd_raw)
            # pattern = r'\{[^{}]*\}'
            # fd_raw = re.findall(pattern, fd_raw)
            fd_dict = ast.literal_eval(fd_raw)
            break
        except Exception as e:
            print('fd generation failed', e)
    FDs[column] = fd_dict
    pass


def code_evaluation_execution(code, val_data, column):
    # 在val_data上运行code,检查是否能修复错误
    # 使用exec()执行code
    exec(code, globals())
    # 检查修复结果
    flag=True
    for error in val_data.values():
        if str(correct(error['dirty_value'])) != error['ground_truth']:
            print(f"expect {error['ground_truth']}, get {correct(error['dirty_value'])}")
            flag = False
    if flag==False:
        print(f'{column} code failed')
        return flag
    # 验证通过，在corrections上执行代码
    mask = detection_human_repaired[column] == 1
    corrections.loc[mask, column] = corrections.loc[mask, column].apply(correct)
    corrections[column] = corrections[column].astype(str)
    print(f'{column} code passed')
    return True


def fd_evaluation_execution(fd:dict, val_data:dict, column:str):
    match = re.search(r"^(.+?)\s*→\s*(.+)$", fd['functional_dependency'])
    if match:
        attr1, attr2 = match.groups()
    else:
        return False
    # 在val_data上验证fd
    # 验证clean_data时，应该将自己排除在外
    if attr1 == 'None':
        return False
    flag = True
    filter_conditions = (detection_human_repaired[attr1] == 0) & (detection_human_repaired[attr2] == 0)

    for idx, error in val_data.items():
        if error['error_type'] == 'clean':
            filter_conditions &= (dirty_data_human_repaired.index != idx)
    
        if attr1 in error['dirty_tuple']:
            valid_rows = dirty_data_human_repaired[attr1] == error['dirty_tuple'][attr1]
        else:
            valid_rows = pd.Series([False] * len(dirty_data_human_repaired))
            
        attr2_values = dirty_data_human_repaired.loc[valid_rows & filter_conditions, attr2]
        
        if not attr2_values.empty:
            corrected_value = attr2_values.mode()[0]
        else:
            corrected_value = error['dirty_value']
        
        if corrected_value != error['ground_truth']:
            print(f"expect {error['ground_truth']}, get {corrected_value}")
            flag = False
            break
    if flag == False:
        print(f'{column} failed')
        return flag
    for idx in detection_human_repaired.index[detection_human_repaired[column] == 1]:
        valid_rows = (dirty_data_human_repaired[attr1] == dirty_data_human_repaired.at[idx, attr1]) & (detection_human_repaired[attr1] == 0) & (detection_human_repaired[attr2] == 0)
        attr2_values = dirty_data_human_repaired.loc[valid_rows, attr2]
        if not attr2_values.empty:
            corrected_value = attr2_values.mode()[0]
        else:
            corrected_value = corrections.at[idx, column]
        corrections.at[idx, column] = corrected_value
    print(f'{column} fd passed')
    return True


def repair_tableau():
    # llm生成的specific examples
    sp_examps = {}
    # 采样得到的代表性错误，d[column][index], 1. dirty_tuple 2. dirty_value 3. ground_truth 4. error_analysis 5. error_type
    rep_error_info = {}
    time_auto_cot_start = time.time()
    retriever_build_time = 0
    print('Generating Auto-CoT...')
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for col_idx, column in enumerate(detection.columns):
            if detection[column].sum() > 0:
                print(column)
                print('Building retriever...')
                retriever_start = time.time()
                retriever, indices, CoE = build_retriever_3(dirty_data,
                                                            detection,
                                                            embeddings_matrix,
                                                            column,
                                                            )
                retriever_build_time += time.time() - retriever_start
                print('Building completed')
                retriever_dict[column] = retriever
                indices_dict[column] = indices
                CoE_dict[column] = CoE
                detection_filtered = detection.iloc[:, indices]
                future = executor.submit(get_auto_cot,
                                         repair_list,
                                         column,
                                         retriever,
                                         CoE,
                                         indices,
                                         detection_filtered,
                                         rep_error_info,
                                         sp_examps
                                         )
                futures.append(future)
        for future in as_completed(futures):
            future.result()
    # 写入specifc examples
    with open(os.path.join(output_path, 'specific_examples.txt'), 'w', encoding='utf-8') as f_output:
        for column, few_shot_specific_str in sp_examps.items():
            f_output.write(f"Column: {column}\n")
            f_output.write(f"Examples:\n{few_shot_specific_str}\n")
            f_output.write("\n" + "="*50 + "\n\n")
    print('Auto-CoT generation completed') 
    time_auto_cot_end = time.time()
    print('time cost: ', time_auto_cot_end - time_auto_cot_start)
    f_time_cost.write(f"Auto-CoT generation time cost: {time_auto_cot_end - time_auto_cot_start}\n")
    print(f"retriever build time: {retriever_build_time}")
    f_time_cost.write(f"retriever build time: {retriever_build_time}\n")
    
    rep_clean_data = sel_clean(human_repair_num)
    # 融合采样得到的代表性错误和正确数据
    rep_data_info = copy.deepcopy(rep_error_info)
    if not rep_clean_data.empty:
        clean_data_integration(rep_clean_data, rep_data_info)

    # 生成代码/函数依赖
    print('Start code generation')
    time_code_generation_start = time.time()
    codes = {}
    fds = {}
    train_data = {}
    val_data = {}
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for column in detection.columns:
            if detection[column].sum() > 0:
                formatting_issue = True
                for error in rep_error_info[column].values():
                    if error['error_type'] != 'clean':
                        if error['error_type'] != 'Formatting Issue':
                            formatting_issue=False
                            break
                train_data[column], val_data[column] = train_val_split(rep_data_info[column])
                if formatting_issue:
                    future = executor.submit(code_generation,
                                            train_data[column],
                                            column,
                                            codes
                                            )
                else:
                    future = executor.submit(FD_generation,
                                            train_data[column],
                                            column,
                                            fds
                                            )
                futures.append(future)
        for future in as_completed(futures):
            future.result()
    with open(os.path.join(output_path, 'codes.txt'), 'w', encoding='utf-8') as f_output:
        for column, code in codes.items():
            f_output.write(f"Column: {column}\n")
            f_output.write(f"Code:\n{code}\n")
            f_output.write("\n" + "="*50 + "\n\n")
    with open(os.path.join(output_path, 'fds.txt'), 'w', encoding='utf-8') as f_output:
        for column, fd in fds.items():
            f_output.write(f"Column: {column}\n")
            f_output.write(f"Fd:\n{json.dumps(fd)}\n")
            f_output.write("\n" + "="*50 + "\n\n")
    time_code_generation_end = time.time()
    print('Code generation completed')
    print('time cost: ', time_code_generation_end - time_code_generation_start)
    f_time_cost.write(f"Code generation time cost: {time_code_generation_end - time_code_generation_start}\n")
    
    # 验证代码&执行代码
    print('Start code evaluation and execution')
    time_code_evaluation_start = time.time()
    for column in codes.keys():
        code_evaluation_execution(codes[column],val_data[column],column)
    for column in fds.keys():
        fd_evaluation_execution(fds[column],val_data[column],column)
    time_code_evaluation_end = time.time()
    print('Code evaluation and execution completed')
    print('time cost: ', time_code_evaluation_end - time_code_evaluation_start)
    f_time_cost.write(f"Code evaluation and execution time cost: {time_code_evaluation_end - time_code_evaluation_start}\n")
    # 在接下来用LLM进行修复时，只需要考虑未被代码修复的元素
    # 对比corrections和dirty_data_human_repaired，找出dirty_data_human_repaired中被corrections修复的元素，然后更新detection_human_repaired，将这些元素对应的detection_human_repaired中的元素置为0
    detection_human_repaired_copy = detection_human_repaired.copy()
    mask = dirty_data_human_repaired != corrections
    detection_human_repaired_copy[mask] = 0
    # 更新检索器
    print('Start updating retriever')
    time_update_retriever_start = time.time()
    for col_idx, column in enumerate(detection_human_repaired.columns):
        if detection_human_repaired[column].sum() > 0:
            update_retriever(column)
    time_update_retriever_end = time.time()
    print('Updating retriever completed')
    print('time cost: ', time_update_retriever_end - time_update_retriever_start)
    f_time_cost.write(f"Updating retriever time cost: {time_update_retriever_end - time_update_retriever_start}\n")
    # 检索
    print('Start retrieving')
    time_retrieving_start = time.time()
    total_time = 0
    retriever_time = 0
    dict_creation_time = 0
    sort_time = 0
    for col_idx, column in enumerate(dirty_data_human_repaired.columns):
        if detection_human_repaired[column].sum() > 0:
            retriever = retriever_dict[column]
            indices = indices_dict[column]
            CoE = CoE_dict[column]
            temp = detection_human_repaired.iloc[:, indices]
            for row_idx in range(len(detection_human_repaired)):
                # 只需要检索未被Function修复的元素，但是排序需要依据原来的Detection
                if detection_human_repaired_copy.at[row_idx, column] == 1:
                    # 在外层循环中转换为NumPy数组
                    column_detection = detection_human_repaired[column].values
                    column_indices = indices_dict[column]
                    column_sums = detection_human_repaired.iloc[:, column_indices].values.sum(axis=1)
                    start_time = time.time()
                    relevant_clean_tuples = ''
                    embeddings_row = embeddings_matrix[row_idx]
                    embeddings_row_filtered = embeddings_row[indices]
                    for i in range(len(embeddings_row_filtered)):
                        if temp.iloc[row_idx, i] == 1:
                            embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
                    embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
                    embeddings_row_united = embeddings_row_filtered.flatten()
                    retriever_start = time.time()
                    relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
                                                                                     k=30)
                    retriever_time += time.time() - retriever_start
                    dict_start = time.time()
                    # 在内层循环中使用NumPy数组
                    relevant_rows_dict_list = [
                        {
                            'page_content': row[0].page_content,
                            'index': idx,
                            'score': round(row[1], 2),
                            'target_column': column_detection[idx],
                            'sum': column_sums[idx]
                        }
                        for row in relevant_rows
                        for idx in [row[0].metadata['index']]
                    ]
                    dict_creation_time += time.time() - dict_start
                    sort_start = time.time()
                    sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
                                                                'sum')
                    sort_time += time.time() - sort_start
                    for row in sorted_relevant_rows_dict_list[:3]:
                        relevant_clean_tuples += row['page_content'] + '\n'
                    if column not in retrieved_tuples:
                        retrieved_tuples[column] = {}
                    retrieved_tuples[column][row_idx] = relevant_clean_tuples
                    total_time += time.time() - start_time
    print(f"Total time: {total_time:.2f} seconds")
    # print(f"Retriever time: {retriever_time:.2f} seconds ({retriever_time / total_time * 100:.2f}%)")
    # print(f"Dict creation time: {dict_creation_time:.2f} seconds ({dict_creation_time / total_time * 100:.2f}%)")
    # print(f"Sort time: {sort_time:.2f} seconds ({sort_time / total_time * 100:.2f}%)")
    # print(
    #     f"Other operations: {(total_time - retriever_time - dict_creation_time - sort_time):.2f} seconds ({(total_time - retriever_time - dict_creation_time - sort_time) / total_time * 100:.2f}%)")
    # time_real_correction_start = time.time()
    print('Retrieval completed')
    time_retrieving_end = time.time()
    print('time cost: ', time_retrieving_end - time_retrieving_start)
    f_time_cost.write(f"Retrieval time cost: {time_retrieving_end - time_retrieving_start}\n")
    # 修复
    print('Start repairing')
    time_repairing_start = time.time()
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for col_idx, column in enumerate(detection_human_repaired.columns):
            # 若列中存在错误
            if detection_human_repaired_copy[column].sum() > 0:
                # 自己首先使用循环组成Input,再将Input传入prompt
                # 遍历列中每个元素
                prompt = ChatPromptTemplate(
                    messages=[
                        SystemMessagePromptTemplate.from_template(sys),
                        HumanMessagePromptTemplate.from_template(human),
                    ],
                    partial_variables={
                        "general_examples": general_examples_str,
                        "specific_examples": sp_examps[column],
                        "format_instructions": parser.get_format_instructions()
                    }
                )
                prompt_dict[column] = prompt
                chain = (
                        prompt
                        | llm
                        | parser
                )
                for row_idx in range(len(detection_human_repaired)):
                    if detection_human_repaired_copy.at[row_idx, column] == 1:
                        # dirty tuple
                            dirty_tuple = dirty_data_human_repaired.iloc[row_idx]
                            dirty_value = dirty_data_human_repaired.at[row_idx, column]
                            # repair the dirty value
                            future = executor.submit(repair_value,
                                                     dirty_tuple,
                                                     column,
                                                     dirty_value,
                                                     row_idx,
                                                     col_idx,
                                                     chain,
                                                     )
                            futures.append(future)
                    # 等待所有的future完成
                    # del retriever
        for future in as_completed(futures):
            future.result()
    time_repairing_end = time.time()
    print('Repairing completed')
    print('time cost: ', time_repairing_end - time_repairing_start)
    f_time_cost.write(f"Repairing time cost: {time_repairing_end - time_repairing_start}\n")


def repair_value(dirty_tuple, column, dirty_value, index_row, index_col, chain):
    global total_tokens
    filtered_tuple = dirty_tuple.iloc[indices_dict[column]]
    filtered_header = [header[i] for i in indices_dict[column]]
    dirty_tuple_filtered_str = format_row(filtered_tuple, filtered_header)
    dirty_value_str = '{' + f'"{column}": "{dirty_value}"' + '}'
    dirty_tuple_json = dirty_tuple.to_dict()
    correction = dirty_value
    result_json = dirty_value_str
    # relevant_clean_tuples = ''
    # embeddings_row = embeddings_matrix[index_row]
    # embeddings_row_filtered = embeddings_row[indices_dict[column]]
    # for i in range(len(embeddings_row_filtered)):
    #     if detection_human_repaired.iloc[:, indices_dict[column]].iloc[index_row, i] == 1:
    #         embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
    # embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE_dict[column][:, np.newaxis])
    # embeddings_row_united = embeddings_row_filtered.flatten()
    # relevant_rows = retriever_dict[column].similarity_search_by_vector(embedding=embeddings_row_united,
    #                                                                    k=3)
    # for row in relevant_rows:
    #     relevant_clean_tuples += row.page_content + '\n'
    relevant_clean_tuples = retrieved_tuples[column][index_row]
    try_num = 0
    while True:
        try:
            promt = prompt_dict[column].invoke({'Dirty_Tuple': dirty_tuple_filtered_str,
                                                'Erroneous_value': dirty_value_str,
                                                'Relevant_clean_tuples': relevant_clean_tuples,
                                                })
            # with open('prompts_correction.txt', 'w', encoding='utf-8') as f3:
            #     f3.write(promt.to_string())
            repair_result = chain.invoke({'Dirty_Tuple': dirty_tuple_filtered_str,
                                          'Erroneous_value': dirty_value_str,
                                          'Relevant_clean_tuples': relevant_clean_tuples,
                                          })
            total_tokens += len(tokenizer.encode(str(promt)))
            total_tokens += len(tokenizer.encode(json.dumps(repair_result)))
            result_json = repair_result
            if result_json['correction'].get(column) is None:
                result_json['correction'][column] = 'null'
            correction = result_json['correction'][column]
            break
        except Exception as e:
            print('ChatModel请求错误', e)
            try_num += 1
            if try_num >= 999:
                break
            continue
    corrections.iloc[index_row, index_col] = str(correction)
    # global num
    # if num < 500 and column_name == 'rate':
    #     print(str(correction))
    #     num = num + 1
    log = {'Index': dirty_tuple_json['index'],
           'Dirty_tuple': format_row(dirty_tuple, header),
           'Dirty_value': dirty_value_str,
           'Relevant_clean_tuples': relevant_clean_tuples,
           'Correction': str(result_json)
           }
    logs.append(log)


def cmp_mark(df_A, df_B):
    # 将不一致元素标为高亮
    df_A.fillna('null', inplace=True)
    df_B.fillna('null', inplace=True)
    # 找出不同的元素
    difference = df_A.ne(df_B)
    # 计算不一致元素的总数
    diff_count = difference.sum().sum()
    # 计算不一致元素的百分比
    total_elements = difference.size
    diff_percent = diff_count / total_elements * 100

    # 新建一个Excel文件
    wb = Workbook()
    ws = wb.active

    # 将B DataFrame的数据填充到工作表中
    for r in dataframe_to_rows(df_B, index=False, header=True):
        ws.append(r)

    # 遍历difference DataFrame，为B中与A不同的单元格添加高亮
    for col in range(difference.shape[1]):
        # 注意：openpyxl 从1开始计数，所以我们需要加2（1表示从第1行开始，另外1用于跳过标题行）
        for row in range(difference.shape[0]):
            if difference.iloc[row, col]:
                cell = ws.cell(row=row + 2, column=col + 1)
                # 设置单元格的填充色为黄色
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 打印不一致元素的个数和百分比
    print(f"不一致元素个数: {diff_count}, 百分比: {diff_percent:.2f}%")
    # 保存为xlsx文件
    filename_with_extension = os.path.basename(dirty_data_path)
    filename_without_extension = os.path.splitext(filename_with_extension)[0]
    save_path = os.path.join(output_path, f'{filename_without_extension}-corrected-marked_{MODEL_NAME}.xlsx')
    wb.save(save_path)


def harmonic_mean(a, b):
    # 调和平均值
    return 2 / (1 / a + 1 / b)


def calc_p_r_f(clean_data, dirty_data, corrected_data):
    # 计算P, R, F
    # corrected_data和dirty_data不一致时，表示模型对该元素进行了修复
    mask_bc = corrected_data != dirty_data
    # corrected_data和clean_data一致时，表示模型对该元素修复成功
    mask_ac = clean_data == corrected_data
    corrected_num = mask_bc.sum().sum()
    final_mask1 = mask_ac & mask_bc
    right_corrected_num1 = final_mask1.sum().sum()
    Precision = right_corrected_num1 / corrected_num
    # clean_data和dirty_data不一致时，表示该元素是错误的
    mask_ab = clean_data != dirty_data
    dirty_num = mask_ab.sum().sum()
    # corrected_data和clean_data一致，表示该错误元素成功修复
    final_mask2 = mask_ab & mask_ac
    right_corrected_num2 = final_mask2.sum().sum()
    Recall = right_corrected_num2 / dirty_num
    F1 = harmonic_mean(Precision, Recall)
    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        print(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}')
        f_output.write(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}\n')


def save_print_logs(logs):
    with open(os.path.join(output_path, 'output.txt'), 'w', encoding='utf-8') as f_output:
        for log in logs:
            print(f"Dirty_tuple:\n{log['Dirty_tuple']}")
            print(f"Dirty_value:\n{log['Dirty_value']}")
            print(f"Relevant_clean_tuples:\n{log['Relevant_clean_tuples']}")
            print(f"Correction:\n{log['Correction']}")
            f_output.write(f"Dirty_tuple:\n{log['Dirty_tuple']}\n")
            f_output.write(f"Dirty_value:\n{log['Dirty_value']}\n")
            f_output.write(f"Relevant_clean_tuples:\n{log['Relevant_clean_tuples']}\n")
            f_output.write(f"Correction:\n{log['Correction']}\n")


def process_dirty_data(dirty_data, detection):
    processed_dirty_data = pd.DataFrame(np.where(detection == 1, 'null', dirty_data), columns=dirty_data.columns)
    return processed_dirty_data


def calc_mi_2(df, target_column):
    # 计算互信息的第二种方式，先筛选后计算
    # Make sure the target_column exists in dataframe
    from joblib import Parallel, delayed
    def calculate_mi_optimized(column, target_column, df):
        if column == target_column:
            return mutual_info_score(df[target_column], df[column])
        else:
            # 计算每个组合的计数，并过滤掉只出现一次的组合
            counts = df.groupby([target_column, column]).size()
            filtered_counts = counts[counts > 1].reset_index()
            
            # 如果过滤后的数据为空，返回0
            if filtered_counts.empty:
                return 0
            
            # 重新构建target和column以计算MI
            target_values = filtered_counts[target_column]
            column_values = filtered_counts[column]
            return mutual_info_score(target_values, column_values)

    # 并行计算所有列的 mutual information
    mutual_info_list = Parallel(n_jobs=-1)(
        delayed(calculate_mi_optimized)(col, target_column, df) for col in df.columns
    )

    # 找到最大 mutual information 值并归一化
    max_mutual_info = max(mutual_info_list) if mutual_info_list else 0
    normalized_mi = [mi / max_mutual_info if max_mutual_info != 0 else 0 for mi in mutual_info_list]
    return normalized_mi


def format_row(row, header):
    s = '{' + ', '.join(f'"{col}": "{val}"' for col, val in zip(header, row)) + '}'
    return s


def format_row_2(value, key, detection_row):
    result = {key[i]: value[i] for i in range(len(value)) if detection_row[i] == 0}
    return json.dumps(result)


def set_dirty_zeros(np_array, df):
    # 将脏数据的句向量置为全0作为惩罚
    l = np_array.shape[2]
    for idx, row in df.iterrows():
        for col, value in row.items():
            if value == 1:
                # Set corresponding vector in numpy array to zeros
                np_array[idx, df.columns.get_loc(col)] = np.zeros(l)
    return np_array


def load_prompts(*file_paths: str) -> tuple:
    return tuple(open(path, 'r', encoding='utf-8').read() for path in file_paths)


def load_examples(file_path: str) -> str:
    with open(file_path, 'r', encoding='utf-8') as file:
        content = ast.literal_eval(file.read())
    return form_examples(content)


class myEmbeddings(Embeddings):
    # 自定义嵌入
    def __init__(self, modelPath):
        self.model = FastSentenceTransformer(modelPath, device="cuda", quantize=True)

    def embed_documents(self, texts: List[str]) -> np.ndarray:
        return self.model.encode(texts, normalize_embeddings=True,batch_size=64,show_progress_bar=True)

    def embed_query(self, text: str) -> np.ndarray:
        return self.model.encode(text, normalize_embeddings=True)


class myEmbeddings_2(Embeddings):
    # 自定义嵌入
    def __init__(self, modelPath):
        self.model = SentenceTransformer(r"E:\LLM-FT-Test\all-MiniLM-L6-v2",
                            backend="onnx",
                            model_kwargs={"file_name": "onnx/model_O4.onnx"},
                           )

    def embed_documents(self, texts: List[str]) -> np.ndarray:
        return self.model.encode(texts, normalize_embeddings=True)

    def embed_query(self, text: str) -> np.ndarray:
        return self.model.encode(text, normalize_embeddings=True)


if __name__ == "__main__":
    total_tokens = 0
    tokenizer = get_tokenizer('qwen-turbo')
    f_time_cost = open('rayyan_time_cost.txt', 'a', encoding='utf-8')
    prompt_dict = {}
    human_repair_num = 10
    # param-csv_file_path
    clean_data_path = 'datasets/rayyan/rayyan_clean.csv'
    dirty_data_path = 'datasets/rayyan/rayyan_dirty.csv'
    detection_path = 'datasets/rayyan/rayyan_dirty_error_detection.csv'
    output_path = get_folder_name('runs_rayyan')
    # param-prompts_path
    system_message_path = 'prompt_templates-old/SystemMessage-2.txt'
    human_message_path = 'prompt_templates-old/HumanMessage.txt'
    general_examples_path = 'prompt_templates-old/examples.txt'
    system_message_auto_cot_path = 'prompt_templates-old/SystemMessage_for_AutoCoT_with_error_type.txt'
    examples_auto_cot_path = 'prompt_templates-old/examples_for_AutoCoT_with_error_type.txt'
    human_message_auto_cot_large_path = 'prompt_templates-old/HumanMessage_for_AutoCoT_large.txt'
    human_message_auto_cot_small_path = 'prompt_templates-old/HumanMessage_for_AutoCoT_small.txt'
    system_message_data_augmentation_path = 'prompt_templates-old/SystemMessage_data_augmentation.txt'
    human_message_data_augmentation_path = 'prompt_templates-old/HumanMessage_data_augmentation.txt'
    system_message_code_generation_path = 'prompt_templates-old/SystemMessage_code_generation.txt'
    human_message_code_generation_path = 'prompt_templates-old/HumanMessage_code_generation.txt'
    system_message_fd_generation_path = 'prompt_templates-old/SystemMessage_fd_generation.txt'
    human_message_fd_generation_path = 'prompt_templates-old/HumanMessage_fd_generation.txt'
    # param-llm
    # The model name, such as gpt-3.5-turbo
    MODEL_NAME = ''
    # The base URL for the API endpoint, such as OPENAI's API endpoint
    OPENAI_API_BASE = ''
    # The API key for authentication
    OPENAI_API_KEY = ''
    # Sampling temperature for response diversity
    TEMPERATURE = 0.5
    # The model name for the AutoCoT version, such as gpt-4o
    MODEL_NAME_auto_cot = ''
    # The base URL for the API endpoint in AutoCoT
    OPENAI_API_BASE_auto_cot = ''
    # The API key for authentication in AutoCoT
    OPENAI_API_KEY_auto_cot = ''
    # Sampling temperature for response generation in AutoCoT
    TEMPERATURE_auto_cot = 0.5
    # Sampling temperature for response generation in data augmentation
    TEMPERATURE_data_augmentation = 0.3
    # param-code_generation
    # The model name for the code generation version, such as gpt-4o
    MODEL_NAME_code_generation = ''
    # The base URL for the API endpoint in code generation
    OPENAI_API_BASE_code_generation = ''
    # The API key for authentication in code generation
    OPENAI_API_KEY_code_generation = ''
    # Sampling temperature for response generation in code generation
    TEMPERATURE_code_generation = 0
    # param-fd_generation
    # The model name for the fd generation version, such as gpt-4o
    MODEL_NAME_fd_generation = ''
    # The base URL for the API endpoint in fd generation
    OPENAI_API_BASE_fd_generation = ''
    # The API key for authentication in fd generation
    OPENAI_API_KEY_fd_generation = ''
    # Sampling temperature for response generation in fd generation
    TEMPERATURE_fd_generation = 0
    # param-embedding
    # The path to the embedding model, embedding models such as sentencebert, word2vec, fasttext, etc.
    EMBEDDING_MODEL_PATH = ''
    #########################################################
    
    
    general_examples_str = load_examples(general_examples_path)
    examples_auto_cot_str = load_examples(examples_auto_cot_path)
    
    sys, human = load_prompts(system_message_path, human_message_path)
    sys_auto_cot, human_auto_cot_large, human_auto_cot_small = load_prompts(
        system_message_auto_cot_path,
        human_message_auto_cot_large_path,
        human_message_auto_cot_small_path
    )
    sys_data_augmentation, human_data_augmentation = load_prompts(system_message_data_augmentation_path, human_message_data_augmentation_path)
    sys_code_generation, human_code_generation = load_prompts(system_message_code_generation_path, human_message_code_generation_path)
    sys_fd_generation, human_fd_generation = load_prompts(system_message_fd_generation_path, human_message_fd_generation_path)
    
    # Initialize the ChatOpenAI instance for standard language model with specified configurations
    llm = ChatOpenAI(
        model_name = MODEL_NAME,  # The model name to be used, e.g., 'gpt-3.5-turbo-0125'
        openai_api_base = OPENAI_API_BASE,  # Base URL for OpenAI API endpoint
        openai_api_key = OPENAI_API_KEY,  # API key for authentication
        temperature = TEMPERATURE  # Sampling temperature for response diversity
    )
    # Initialize the ChatOpenAI instance for an AutoCoT-specific language model with separate configurations
    llm_auto_cot = ChatOpenAI(
        model_name = MODEL_NAME_auto_cot,  # The model name for the AutoCoT version
        openai_api_base = OPENAI_API_BASE_auto_cot,  # Base URL for OpenAI API endpoint in AutoCoT
        openai_api_key = OPENAI_API_KEY_auto_cot,  # API key specific to the AutoCoT configuration
        temperature = TEMPERATURE_auto_cot  # Sampling temperature for response generation in AutoCoT
    )
    # llm_data_augmentation = ChatOpenAI(
    #     model_name = MODEL_NAME_data_augmentation,  # The model name for the AutoCoT version
    #     openai_api_base = OPENAI_API_BASE_data_augmentation,  # Base URL for OpenAI API endpoint in AutoCoT
    #     openai_api_key = OPENAI_API_KEY_data_augmentation,  # API key specific to the AutoCoT configuration
    #     temperature = TEMPERATURE_data_augmentation  # Sampling temperature for response generation in AutoCoT
    # )
    llm_code_generation = ChatOpenAI(
        model_name = MODEL_NAME_code_generation,  # The model name for the AutoCoT version
        openai_api_base = OPENAI_API_BASE_code_generation,  # Base URL for OpenAI API endpoint in AutoCoT
        openai_api_key = OPENAI_API_KEY_code_generation,  # API key specific to the AutoCoT configuration
        temperature = TEMPERATURE_code_generation  # Sampling temperature for response generation in AutoCoT
    )
    llm_fd_generation = ChatOpenAI(
        model_name = MODEL_NAME_fd_generation,  
        openai_api_base = OPENAI_API_BASE_fd_generation,
        openai_api_key = OPENAI_API_KEY_fd_generation,  
        temperature = TEMPERATURE_fd_generation
    )
    # the parser for the output of the llm
    parser = JsonOutputParser(pydantic_object=Output)
    # load the clean data and dirty data
    clean_data = pd.read_csv(clean_data_path, dtype=str, encoding='utf-8')
    clean_data.fillna('null', inplace=True)
    dirty_data = pd.read_csv(dirty_data_path, dtype=str, encoding='utf-8')
    dirty_data.fillna('null', inplace=True)
    
    row_count, column_count = dirty_data.shape
    header = dirty_data.columns.tolist()
    # load the detection result
    detection = pd.read_csv(detection_path)
    print(detection.sum().sum())
    # initialize the logs
    logs = []
    # initialize the corrections
    corrections = dirty_data.copy()
    start_time = time.time()
    # get the elements list
    elements_list = dirty_data.values.flatten().tolist()
    print('embedding...')
    emb_start_time = time.time()
    embeddingModel = myEmbeddings(EMBEDDING_MODEL_PATH)
    embeddings = embeddingModel.embed_documents(elements_list)
    emb_end_time = time.time()
    embedding_dimension = len(embeddings[0])
    embeddings_matrix = embeddings.reshape(row_count, column_count, embedding_dimension)
    print('embedding done')
    print(f"embedding time cost {emb_end_time-emb_start_time}")
    f_time_cost.write(f"embedding time cost {emb_end_time-emb_start_time}\n")
    # select the user annotation candidates
    select_start_time = time.time()
    repair_list = select_repair_candidates(embeddings_matrix, detection, human_repair_num)
    select_end_time = time.time()
    print('select done')
    print(f"select time cost {select_end_time-select_start_time}")
    f_time_cost.write(f"select time cost {select_end_time-select_start_time}\n")
    print(repair_list)
    # repair the dirty data
    dirty_data_human_repaired = dirty_data.copy()
    dirty_data_human_repaired.iloc[repair_list] = clean_data.iloc[repair_list]
    detection_human_repaired = detection.copy()
    detection_human_repaired.iloc[repair_list] = 0
    corrections.iloc[repair_list] = clean_data.iloc[repair_list]
    dirty_data_only_repaired = clean_data.iloc[repair_list]
    elements_list_only_repaired = dirty_data_only_repaired.values.flatten().tolist()
    embeddings_only_repaired = embeddingModel.embed_documents(elements_list_only_repaired)
    embeddings_matrix_only_repaired = embeddings_only_repaired.reshape(human_repair_num, column_count,
                                                                       embedding_dimension)
    retriever_dict = {}
    indices_dict = {}
    CoE_dict = {}
    retrieved_tuples = {}
    repair_tableau()
    end_time = time.time()
    logs = sorted(logs, key=lambda x: int(x['Index']))
    # save_print_logs(logs)
    corrections.to_csv(os.path.join(output_path, 'corrections.csv'), encoding='utf-8',index=False)
    cmp_mark(clean_data, corrections)
    calc_p_r_f(clean_data, dirty_data_human_repaired, corrections)
    execution_time = end_time - start_time
    print(f"execution time: {execution_time} seconds")
    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        f_output.write(f"execution time: {execution_time} seconds\n")
    print(f"total tokens: {total_tokens}")
    f_time_cost.close()